import csv
import os
import re
import sys
import shutil
import unicodedata
from datetime import datetime
from tkinter import filedialog, messagebox

import openpyxl

# テンプレートは テンプレート/注文書/ サブフォルダに配置
TEMPLATE_SUBFOLDER = '発注書'

# pairs: (品名列, 商品コード列, 個数列) ※コード列が None の場合は名前照合
# キー：ファイル名に含まれるキーワード（部分一致）
TEMPLATE_CONFIGS = {
    '味と暮らし': {'pairs': [('B','C','D'), ('F','G','H')], 'rows': (10, 31)},
    'カレーフェスタ': {'pairs': [('B','C','D'), ('F','G','H')], 'rows': (10, 32)},
    'F15':        {'pairs': [('A','C','D'), ('E','G','H')], 'rows': (11, 26)},
    '若竹会':     {'pairs': [('B',None,'J')],               'rows': (7,  19)},
}


def normalize(s: str) -> str:
    """全角→半角・空白正規化（表記ゆれ吸収）"""
    return ' '.join(unicodedata.normalize('NFKC', s).split())


def get_template_dir() -> str:
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, 'テンプレート', TEMPLATE_SUBFOLDER)


def read_delivery_csv(filepath: str) -> tuple:
    """注文書用CSV → ({商品コード: 数量}, {正規化商品名: 数量})"""
    qty_by_code = {}
    qty_by_name = {}
    with open(filepath, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [k.lstrip('﻿').strip() for k in reader.fieldnames]
        cols = set(reader.fieldnames or [])
        dept_cols = {'部署名', '氏名', '注文番号'}
        if dept_cols.issubset(cols):
            raise ValueError(
                '部署別CSVが選択されています。\n'
                'WordPressの「注文書用に出力」でダウンロードしたCSVを選択してください。'
            )
        if '単価(税込)' in cols or '合計金額(税込)' in cols:
            raise ValueError(
                '納品書用CSVが選択されています。\n'
                'WordPressの「注文書用に出力」でダウンロードしたCSVを選択してください。'
            )
        for row in reader:
            name    = row.get('商品名', '').strip()
            code    = row.get('商品コード', '').strip()
            qty_str = row.get('合計数量', '0').strip().replace(',', '')
            if not name:
                continue
            try:
                qty = int(float(qty_str))
            except ValueError:
                qty = 0
            if code:
                qty_by_code[code] = qty_by_code.get(code, 0) + qty
            qty_by_name[normalize(name)] = qty_by_name.get(normalize(name), 0) + qty
    return qty_by_code, qty_by_name


def fill_template(template_path: str, qty_by_code: dict, qty_by_name: dict, output_path: str) -> tuple:
    """テンプレートに個数を書き込み、(マッチ数, 総商品数) を返す"""
    fname = os.path.basename(template_path)
    config = next((v for k, v in TEMPLATE_CONFIGS.items() if k in fname), None)
    if not config:
        return 0, 0

    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    matched = 0
    total = 0
    min_row, max_row = config['rows']

    for name_col, code_col, qty_col in config['pairs']:
        for r in range(min_row, max_row + 1):
            name_val = ws[f'{name_col}{r}'].value
            if name_val is None:
                continue
            total += 1

            qty = None
            # 商品コードで照合
            if code_col:
                code_val = ws[f'{code_col}{r}'].value
                if code_val:
                    qty = qty_by_code.get(str(code_val).strip())
            # コードで見つからなければ商品名で照合（完全一致→スペース無視→前方一致）
            if qty is None:
                tmpl = normalize(str(name_val))
                qty = qty_by_name.get(tmpl)
            if qty is None:
                tmpl_ns = tmpl.replace(' ', '')
                for csv_name, v in qty_by_name.items():
                    if csv_name.replace(' ', '') == tmpl_ns:
                        qty = v
                        break
            if qty is None:
                for csv_name, v in qty_by_name.items():
                    if csv_name.startswith(tmpl) and len(csv_name) > len(tmpl) and csv_name[len(tmpl)] == ' ':
                        qty = v
                        break

            if qty is not None:
                ws[f'{qty_col}{r}'] = qty
                matched += 1

    wb.save(output_path)

    # 設定された行範囲を超えて商品が存在しないか確認
    # ※コード列がある場合、コードが空の行はフッター行とみなしスキップ
    over_rows = []
    for name_col, code_col, qty_col in config['pairs']:
        for r in range(max_row + 1, max_row + 20):
            name_val = ws[f'{name_col}{r}'].value
            if not name_val:
                continue
            if code_col and ws[f'{code_col}{r}'].value is None:
                continue
            over_rows.append(str(name_val))

    warning = None
    if over_rows:
        warning = (
            f'【{fname}】\n'
            f'設定行数（{max_row}行）を超えた商品が{len(over_rows)}件あります。\n'
            f'担当者にテンプレート設定の更新を依頼してください。\n'
            f'（対象例: {over_rows[0]}）'
        )

    return matched, total, warning


def run() -> None:
    template_dir = get_template_dir()

    csv_path = filedialog.askopenfilename(
        title='注文書用CSVを選択してください',
        filetypes=[('CSVファイル', '*.csv'), ('すべてのファイル', '*.*')],
    )
    if not csv_path:
        return

    try:
        qty_by_code, qty_by_name = read_delivery_csv(csv_path)
    except Exception as e:
        messagebox.showerror('エラー', f'CSV読み込みエラー:\n{e}')
        return

    if not qty_by_name:
        messagebox.showerror('エラー', 'CSVにデータがありません')
        return

    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_dir = os.path.dirname(csv_path)
    output_paths = []
    total_matched = 0
    total_items = 0

    all_templates = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    warnings = []
    for fname in all_templates:
        template_path = os.path.join(template_dir, fname)
        stem = os.path.splitext(fname)[0]
        out_path = os.path.join(base_dir, f'{stem}_{now}.xlsx')
        try:
            matched, items, warning = fill_template(template_path, qty_by_code, qty_by_name, out_path)
            total_matched += matched
            total_items += items
            if items > 0:
                output_paths.append(out_path)
            if warning:
                warnings.append(warning)
        except Exception as e:
            messagebox.showerror('エラー', f'{fname} の処理中にエラー:\n{e}')
            return

    if warnings:
        messagebox.showwarning('行数オーバー', '\n\n'.join(warnings))

    if not output_paths:
        messagebox.showerror('エラー', 'テンプレートが見つかりませんでした\nテンプレート/注文書/ フォルダを確認してください')
        return

    messagebox.showinfo(
        '完了',
        f'{len(output_paths)}件の注文書を出力しました\n'
        f'個数入力: {total_matched} / {total_items} 商品'
    )
    for path in output_paths:
        os.startfile(path)


if __name__ == '__main__':
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    run()
    root.destroy()
