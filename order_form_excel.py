"""
order_form_excel.py
PDFの発注書テンプレートをExcel形式で生成するスクリプト
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


FONT = "游ゴシック"
GRAY = "C0C0C0"
DARK_GRAY = "808080"


# ── helpers ─────────────────────────────────────────────────────────────────

def _thin():
    return Side(border_style="thin", color="000000")

def _medium():
    return Side(border_style="medium", color="000000")

def _all_thin():
    t = _thin()
    return Border(top=t, bottom=t, left=t, right=t)

def _s(ws, row, col, value="", *, bold=False, sz=9, ha="center", va="center",
       wrap=True, bdr=True, bg=None, fc="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=sz, bold=bold, color=fc)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr:
        c.border = _all_thin()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _m(ws, r1, c1, r2, c2, value="", *, bold=False, sz=9, ha="center",
       va="center", wrap=True, bdr=True, bg=None, fc="000000"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(name=FONT, size=sz, bold=bold, color=fc)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr:
        c.border = _all_thin()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _rh(ws, row, height):
    ws.row_dimensions[row].height = height


# ── product data ─────────────────────────────────────────────────────────────

PRODUCTS_P1 = [
    (1,  "大阪ぐるめすうぷ肉すい",                                   196),
    (2,  "香川県名物ご当地グルメ 薫る骨付鳥 わか",                  265),
    (3,  "静岡おでん 365g",                                           623),
    (4,  "球美のら一油",                                               264),
    (5,  "生しぼりしょうが入り甘酒",                                  501),
    (6,  "殿様するめ",                                                 200),
    (7,  "デーツチョコ 2袋セット",                                    388),
    (8,  "河京の喜多方ラーメン 5食ミックス",                         10),
    (9,  "中国料理「大笑」四川風 白ごま担々麺",                       3),
    (10, "熊本ラーメン繁盛店 もっこす辛とんこつラーメン5食",         56),
    (11, "ご当地ラーメン紀行",                                         79),
    (12, "味噌煮込みうどん 5食",                                       23),
    (13, "田舎造り 信州そば",                                          1),
    (14, "国産味付うずらたまご",                                       239),
    (15, "ラーメン屋さんのチャーシュー 炊き込みご飯の素(2個セット)", 280),
    (16, "北海道産 ほたてめし",                                        293),
    (17, "明太いわし&いわし甘露煮詰め合せ",                          208),
    (18, "やま柿",                                                     300),
    (19, "スティックフィナンシェ しっとりチョコ",                    360),
    (20, "素材たっぷり さつまいもケーキ",                             343),
    (21, "ひととえ こがね芋",                                          329),
    (22, "やきいも大福",                                               339),
    (23, "お徳用 芋キャラミルク黒糖バナナ",                           338),
    (24, "丹波黒大豆入りおかき(TO-B)",                                396),
    (25, "あんころ餅",                                                 390),
    (26, "紀州南部の南高梅(はちみつ)",                                290),
    (27, "ブラックペッパー焼チーズ&明太子焼きチーズサンドセット",   210),
    (28, "めちゃうまラー油あられ",                                    337),
    (29, "竹炭豆 直送便",                                              398),
    (30, "お徳用 生姜黒糖",                                            391),
    (31, "ドリップバッグ スペシャルブレンド",                         552),
    (32, "トン 素焼きナッツ 塩無添加ミックスナッツ",                  216),
    (33, "さかなっつハイ！ 10g×30P",                                 340),
    (34, "健康菓集 ナッツ&フルーツ",                                  202),
    (35, "網焼き 焼き小あじ",                                          218),
    (36, "ヨーグルトレーズン 9g×25袋",                               314),
    (37, "日高昆布",                                                   257),
    (38, "兵庫県産 丹波黒豆",                                          254),
    (39, "SABA 鰯ふりだし",                                            288),
    (40, "北海昆布セット",                                             236),
    (41, "野菜とたまごのスープ",                                       197),
    (42, "いつものおみそ汁6種セット ※受注停止",                      171),
    (43, "毎日のおみそ汁30P",                                          154),
    (44, "北海道コーンスープ",                                         158),
    (45, "スープ6種お楽しみ袋",                                        199),
]

PRODUCTS_P3 = [
    (1,  "信州味噌らあめん 寸八",                                      1601),
    (2,  "名古屋塩らーめん 徳川町 如水",                             1602),
    (3,  "飛騨高山中華そば 麺屋 しらかわ",                           1603),
    (4,  "大分 Furari",                                                1604),
    (5,  "秋田味噌チャンポン チャイナタウン",                         1605),
    (6,  "喜多方らーめん うめハ",                                     1606),
    (7,  "山梨 蓬来軒 支那そば",                                      1607),
    (8,  "喜多方プレミアムラーメン 厚み",                             1608),
    (9,  "喜鈴 黄金醤油ラーメン",                                     1609),
    (10, "山形鶴岡市の名店 中華そば処 琴平荘",                       1610),
    (11, "にいがた王選 長岡生姜醤油ラーメン",                         1611),
    (12, "麺家いろは 富山ブラック黒醤油ラーメン",                    1612),
    (13, "横浜戸塚 支那そばや",                                       1613),
    (14, "和歌山県 お茶の玉林園 てんかけラーメン",                   1614),
    (15, "尾道ラーメン 一丁",                                         1615),
    (16, "鳥取・米子牛骨ラーメン 満州味",                             1616),
    (17, "人形町「玉ひで」監修 軍鶏だし塩らーめん",                  1617),
    (18, "鉢ノ蕎葉",                                                  1618),
    (19, "島根ラーメン かみあり製麺 しじみ塩らーめん",               1619),
    (20, "岡山 小紫",                                                  1620),
    (21, "らーめん堂 仙台っ子",                                       1621),
    (22, "横浜 家系総本山 吉村家",                                    1622),
    (23, "京都ラーメン 無鉄砲",                                       1623),
    (24, "博多ラーメン だるま",                                       1624),
    (25, "博多 一双(4食)",                                            1625),
    (26, "福岡 石田一龍",                                             1626),
    (27, "鹿児島ラーメン 豚とろ(2食)",                               1627),
    (28, "鹿児島 ざぼんラーメン",                                     1628),
    (29, "新京 ベストコンディションラーメン",                         1629),
    (30, "新潟海鮮ラーメン2種詰合せ 新之助 紅ずわいがに・南蛮えびラーメン", 1630),
    (31, "八劔ROCK 人生餃子 皿台湾 汁なし台湾ラーメン",             1631),
    (32, "鈴鹿 門鰓 カレーらーめん",                                 1632),
    (33, "中華蕎麦 とみ田 つけめん",                                 1633),
    (34, "大阪らーめん 颯人",                                         1634),
    (35, "札幌らーめん 四代目いちまる",                               1635),
    (36, "札幌のらーめん 吉山商店",                                   1636),
    (37, "金沢 味噌ラーメン専門店 大河",                             1637),
    (38, "中国料理「天笑」四川風汁なし担々麺",                       1638),
    (39, "東京 創作麺工房 鳴龍 担担麺",                              1639),
    (40, "勝浦式 坦々麺 江ざわ",                                      1640),
    (41, "福島ラーメン うろた・くるん 食べくらべ",                   1641),
    (42, "北海道四大ラーメン (桑名・青葉・一文字・河むら)",          1642),
    (43, "自家製 厚切りチャーシューメンマ 4セット",                  1643),
    (44, "もうたま卵(しょうゆ味)",                                   1644),
    (45, "ガリチップス(やみつきスパイス味)",                         1645),
]

PRODUCTS_F15_LEFT = [
    ("ビーフカレーインド風",                              1148, 103),
    ("ビーフカレーロイヤルプレステージ",                  1809, 104),
    ("ココナッツミルク香るグリーンカレー【チキン】",      1148, 113),
    ("スパイスにこだわったブラックカレー【ポーク】",      1148, 114),
    ("ビーフ&ベジタブルハイブレンド",                    1148, 110),
    ("ビーフカレーデラックスシルバーラベル",              1148, 105),
    ("ハヤシビーフ",                                      1148, 106),
    ("ひとくちカレー",                                    786,  111),
    ("カレーうどんソース金鯱",                            1148, 122),
    ("減塩 ビーフカレー",                                 1700, 107),
    ("減塩 野菜カレー",                                   1700, 108),
    ("まゆまゆ Hokkori 玄米ごはん ガバオライス",         1447, 142),
    ("まゆまゆ Hokkori 玄米ごはん とりごはん",           1447, 143),
    ("まゆまゆ Hokkori 玄米ごはん 五目カレーごはん",     1447, 144),
    ("まぜこみご飯の素 五目ひじき",                       1133, 145),
    ("まぜこみご飯の素 牛ごぼう",                         1133, 148),
    ("まぜこみご飯の素 かしわめし",                       1133, 141),
]

PRODUCTS_F15_RIGHT = [
    ("減脂プーアール茶",                 983, 525),
    ("ノンカフェイン ルイボスティー",   944, 508),
    ("杜仲茶",                            983, 514),
    ("甜茶",                              983, 517),
    ("ヘルシーどくだみ茶",               983, 519),
    ("はと麦茶",                          928, 522),
    ("ジャスミン茶",                      983, 533),
    ("(お徳用)煎茶ティーバッグ",         850, 551),
    ("国内産原料100%使用 なたまめ茶",   691, 507),
]


# ── form builders ─────────────────────────────────────────────────────────

def _setup_cols_standard(ws):
    """カタログNo./品名/商品コード/注文数 × 左右"""
    ws.column_dimensions["A"].width = 5.5
    ws.column_dimensions["B"].width = 24.0
    ws.column_dimensions["C"].width = 7.5
    ws.column_dimensions["D"].width = 7.0
    ws.column_dimensions["E"].width = 5.5
    ws.column_dimensions["F"].width = 24.0
    ws.column_dimensions["G"].width = 7.5
    ws.column_dimensions["H"].width = 7.0


def _build_header_standard(ws, title_line1, title_line2):
    """共通ヘッダーを構築して次の空き行番号を返す"""
    # ── タイトル行 (1-2) ────────────────────────────────────
    _rh(ws, 1, 24)
    _rh(ws, 2, 20)
    _m(ws, 1, 1, 2, 4,
       title_line1 + "\n" + title_line2,
       bold=True, sz=14, ha="left", va="center", bdr=False)
    _m(ws, 1, 5, 2, 8,
       "㈱ユニオンサービス\n本社事務所　名古屋市南区芝町189番地\n〒457-0023　TEL.(052)824-4151",
       sz=8, ha="right", va="center", bdr=False)

    # ── 注文日 (3) ────────────────────────────────────────────
    _rh(ws, 3, 18)
    _m(ws, 3, 1, 3, 8, "注文日　　　月　　日　（　　　　　　　）",
       ha="left", va="center")

    # ── お得意先 / 団体名 (4) ─────────────────────────────────
    _rh(ws, 4, 22)
    _m(ws, 4, 1, 4, 1, "お得意先\nコードNO.", sz=8)
    _m(ws, 4, 2, 4, 3, "")
    _m(ws, 4, 4, 4, 4, "団体名", sz=8)
    _m(ws, 4, 5, 4, 7, "")
    _s(ws, 4, 8, "様", sz=9)

    # ── ご住所 (5) ────────────────────────────────────────────
    _rh(ws, 5, 18)
    _s(ws, 5, 1, "ご住所", sz=8)
    _s(ws, 5, 2, "〒", sz=9, ha="left")
    _m(ws, 5, 3, 5, 4, "")
    _m(ws, 5, 5, 5, 8, "")

    # ── 都道府県 / 市郡区 (6) ────────────────────────────────
    _rh(ws, 6, 18)
    _s(ws, 6, 1, "")
    _m(ws, 6, 2, 6, 2, "都道\n府県", sz=7)
    _s(ws, 6, 3, "")
    _m(ws, 6, 4, 6, 4, "市郡\n区", sz=7)
    _m(ws, 6, 5, 6, 8, "")

    # ── 電話番号 / FAX番号 (7) ────────────────────────────────
    _rh(ws, 7, 18)
    _s(ws, 7, 1, "電話番号", sz=8)
    _m(ws, 7, 2, 7, 3, "")
    _s(ws, 7, 4, "")
    _s(ws, 7, 5, "FAX番号", sz=8)
    _m(ws, 7, 6, 7, 8, "")

    # ── ご担当者様 (8) ────────────────────────────────────────
    _rh(ws, 8, 18)
    _m(ws, 8, 1, 8, 1, "ご担当者様\n氏名", sz=8)
    _m(ws, 8, 2, 8, 3, "")
    _s(ws, 8, 4, "")
    _m(ws, 8, 5, 8, 5, "ご担当者様\n連絡先", sz=8)
    _m(ws, 8, 6, 8, 8, "")

    return 9  # 次の行番号


def _build_product_table_standard(ws, start_row, products):
    """
    商品一覧テーブルを構築。
    左列: products[:23]、右列: products[23:]
    """
    # カラムヘッダー
    _rh(ws, start_row, 22)
    for col, label in [(1, "カタログ\nNo."), (2, "品　名"),
                       (3, "商品\nコード"), (4, "注文数"),
                       (5, "カタログ\nNo."), (6, "品　名"),
                       (7, "商品\nコード"), (8, "注文数")]:
        _s(ws, start_row, col, label, bold=True, sz=8, bg=GRAY)

    left  = products[:23]
    right = products[23:]
    data_start = start_row + 1

    for i in range(23):
        r = data_start + i
        _rh(ws, r, 28)

        if i < len(left):
            no, name, code = left[i]
            _s(ws, r, 1, no,   sz=9)
            _s(ws, r, 2, name, sz=9, ha="left")
            _s(ws, r, 3, code, sz=9)
            _s(ws, r, 4, "",   sz=9)
        else:
            for c in range(1, 5):
                _s(ws, r, c, "")

        if i < len(right):
            no, name, code = right[i]
            _s(ws, r, 5, no,   sz=9)
            _s(ws, r, 6, name, sz=9, ha="left")
            _s(ws, r, 7, code, sz=9)
            _s(ws, r, 8, "",   sz=9)
        else:
            for c in range(5, 9):
                _s(ws, r, c, "")

    # 小計 / 合計点数
    sub_row = data_start + 23
    _rh(ws, sub_row, 20)
    _m(ws, sub_row, 1, sub_row, 3, "小　計", bold=True, sz=9, bg=GRAY)
    _s(ws, sub_row, 4, "", bg=GRAY)
    _m(ws, sub_row, 5, sub_row, 7, "合計点数", bold=True, sz=9, bg=DARK_GRAY, fc="FFFFFF")
    _s(ws, sub_row, 8, "", bg=DARK_GRAY)

    return sub_row + 1  # 次の空き行


def _build_footer_standard(ws, start_row, fax_number="052-824-4152"):
    """備考欄・FAX番号フッターを構築"""
    r = start_row
    _rh(ws, r, 14)
    _m(ws, r, 1, r, 8,
       "※お届け先、到着日などで指定の有る場合は、備考の欄に明記して下さい。",
       sz=8, ha="left", bdr=False)

    r += 1
    note_start = r
    _rh(ws, r, 16)
    _s(ws, r, 1, "備考", sz=8, ha="left")
    _m(ws, r, 2, r + 3, 8, "")
    for rr in range(r + 1, r + 4):
        _rh(ws, rr, 16)
        _s(ws, rr, 1, "")

    r += 4
    _rh(ws, r, 36)
    _m(ws, r, 1, r, 2,
       "商品注文専用\nFAX", sz=9, ha="left", va="center")
    _m(ws, r, 3, r, 8,
       fax_number, bold=True, sz=22, ha="center", va="center")

    ws.page_setup.paperSize = 9           # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:H{r}"


# ── form 1: 味と暮らしの特選街 ────────────────────────────────────────────

def create_form1(ws):
    ws.title = "味と暮らしの特選街"
    _setup_cols_standard(ws)
    next_row = _build_header_standard(
        ws,
        "2025年冬季",
        "味と暮らしの特選街　発注書",
    )
    next_row = _build_product_table_standard(ws, next_row, PRODUCTS_P1)
    _build_footer_standard(ws, next_row, "052-824-4152")


# ── form 3: 日本全国ラーメン専用発注書 ───────────────────────────────────

def create_form3(ws):
    ws.title = "日本全国ラーメン"
    _setup_cols_standard(ws)
    next_row = _build_header_standard(
        ws,
        "2025年",
        "日本全国ラーメン専用発注書",
    )
    next_row = _build_product_table_standard(ws, next_row, PRODUCTS_P3)
    _build_footer_standard(ws, next_row, "052-824-4152")


# ── form 2: 非加盟学童 F15 ────────────────────────────────────────────────

def _setup_cols_f15(ws):
    ws.column_dimensions["A"].width = 24.0
    ws.column_dimensions["B"].width = 10.0
    ws.column_dimensions["C"].width = 7.0
    ws.column_dimensions["D"].width = 7.0
    ws.column_dimensions["E"].width = 24.0
    ws.column_dimensions["F"].width = 10.0
    ws.column_dimensions["G"].width = 7.0
    ws.column_dimensions["H"].width = 7.0


def create_form2(ws):
    ws.title = "非加盟学童F15"
    _setup_cols_f15(ws)

    # ── タイトル (1-2) ────────────────────────────────────────
    _rh(ws, 1, 22)
    _rh(ws, 2, 18)
    _m(ws, 1, 1, 2, 4,
       "【非加盟学童】\n発注書",
       bold=True, sz=12, ha="left", va="center", bdr=False)
    _m(ws, 1, 5, 2, 8,
       "FAX　052－821－8331",
       bold=True, sz=16, ha="right", va="center", bdr=False)

    # ── 注文日 (3) ────────────────────────────────────────────
    _rh(ws, 3, 18)
    _m(ws, 3, 5, 3, 8,
       "注文日　　　月　　日　（　　　　　）",
       ha="right", va="center", bdr=False)

    # ── コードNO. (4) ─────────────────────────────────────────
    _rh(ws, 4, 18)
    _s(ws, 4, 1, "コードNO.", sz=8, ha="left")
    _m(ws, 4, 2, 4, 4, "")
    _m(ws, 4, 5, 4, 8,
       "※「注文日」「コードNO.」「ご担当名」のご記入をお願いします。",
       sz=8, ha="left", bdr=False)

    # ── 作業所名 (5) ─────────────────────────────────────────
    _rh(ws, 5, 18)
    _s(ws, 5, 1, "作業所名", sz=8, ha="left")
    _m(ws, 5, 2, 5, 8, "")

    # ── ご住所 (6) ────────────────────────────────────────────
    _rh(ws, 6, 18)
    _m(ws, 6, 1, 6, 1, "ご住所", sz=8, ha="left")
    _s(ws, 6, 2, "〒", ha="left")
    _m(ws, 6, 3, 6, 8, "")

    # ── 都道府県 (7) ─────────────────────────────────────────
    _rh(ws, 7, 18)
    _s(ws, 7, 1, "")
    _m(ws, 7, 2, 7, 2, "都道\n府県", sz=7)
    _m(ws, 7, 3, 7, 8, "")

    # ── TEL/FAX/ご担当者 (8) ──────────────────────────────────
    _rh(ws, 8, 18)
    _s(ws, 8, 1, "TEL", sz=9, ha="left")
    _s(ws, 8, 2, "FAX", sz=9, ha="left")
    _m(ws, 8, 3, 8, 4, "")
    _s(ws, 8, 5, "ご担当者", sz=8)
    _m(ws, 8, 6, 8, 6, "ご担当者\n連絡先", sz=8)
    _m(ws, 8, 7, 8, 8, "")

    # ── 区分見出し (9) ────────────────────────────────────────
    _rh(ws, 9, 20)
    _m(ws, 9, 1, 9, 4, "F15", bold=True, sz=12, bg=GRAY)
    _m(ws, 9, 5, 9, 8, "F15", bold=True, sz=12, bg=GRAY)

    # ── 商品ヘッダー (10) ─────────────────────────────────────
    _rh(ws, 10, 20)
    for col, label in [(1, "商　品　名"), (2, "卸価格\n(税抜)"),
                       (3, "商品\nコード"), (4, "数量"),
                       (5, "商　品　名"), (6, "卸価格\n(税抜)"),
                       (7, "商品\nコード"), (8, "数量")]:
        _s(ws, 10, col, label, bold=True, sz=8, bg=GRAY)

    # ── 商品行 ────────────────────────────────────────────────
    max_rows = max(len(PRODUCTS_F15_LEFT), len(PRODUCTS_F15_RIGHT))
    data_start = 11
    for i in range(max_rows):
        r = data_start + i
        _rh(ws, r, 28)

        if i < len(PRODUCTS_F15_LEFT):
            name, price, code = PRODUCTS_F15_LEFT[i]
            _s(ws, r, 1, name,          sz=9, ha="left")
            _s(ws, r, 2, f"{price:,}円", sz=9)
            _s(ws, r, 3, code,          sz=9)
            _s(ws, r, 4, "",            sz=9)
        else:
            for c in range(1, 5):
                _s(ws, r, c, "")

        if i < len(PRODUCTS_F15_RIGHT):
            name, price, code = PRODUCTS_F15_RIGHT[i]
            _s(ws, r, 5, name,          sz=9, ha="left")
            _s(ws, r, 6, f"{price:,}円", sz=9)
            _s(ws, r, 7, code,          sz=9)
            _s(ws, r, 8, "",            sz=9)
        else:
            for c in range(5, 9):
                _s(ws, r, c, "")

    # 合計数量
    total_row = data_start + max_rows
    _rh(ws, total_row, 20)
    _m(ws, total_row, 1, total_row, 7, "合計数量", bold=True, sz=9, bg=DARK_GRAY, fc="FFFFFF")
    _s(ws, total_row, 8, "", bg=DARK_GRAY)

    # 注意書き
    note_r = total_row + 1
    _rh(ws, note_r, 14)
    _m(ws, note_r, 1, note_r, 8,
       "※商品の注文金額合計が卸価格で10,000円(税抜)未満の場合は送料600円(税抜)が別途必要になります。",
       sz=8, ha="left", bdr=False)

    # フッター
    fax_r = note_r + 1
    _rh(ws, fax_r, 36)
    _m(ws, fax_r, 1, fax_r, 4,
       "㈱ユニオンサービス\nFAX:052-821-8331\n本社事務所〒457-0023名古屋市南区芝町189番地/TEL052-822-2261",
       sz=7, ha="left", va="center")
    _m(ws, fax_r, 5, fax_r, 8, "", bdr=False)

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:H{fax_r}"


# ── エントリーポイント ────────────────────────────────────────────────────

def get_template_dir():
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "テンプレート")


def run():
    """3種類の発注書Excelテンプレートを生成する"""
    template_dir = get_template_dir()
    os.makedirs(template_dir, exist_ok=True)

    outputs = []

    # ── 1. 味と暮らしの特選街 ─────────────────────────────────
    wb1 = openpyxl.Workbook()
    create_form1(wb1.active)
    path1 = os.path.join(template_dir, "発注書_味と暮らしの特選街.xlsx")
    wb1.save(path1)
    outputs.append(path1)

    # ── 2. 非加盟学童F15 ──────────────────────────────────────
    wb2 = openpyxl.Workbook()
    create_form2(wb2.active)
    path2 = os.path.join(template_dir, "発注書_非加盟学童F15.xlsx")
    wb2.save(path2)
    outputs.append(path2)

    # ── 3. 日本全国ラーメン ───────────────────────────────────
    wb3 = openpyxl.Workbook()
    create_form3(wb3.active)
    path3 = os.path.join(template_dir, "発注書_日本全国ラーメン.xlsx")
    wb3.save(path3)
    outputs.append(path3)

    return outputs


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    try:
        files = run()
        msg = "以下のファイルを生成しました:\n\n" + "\n".join(
            os.path.basename(f) for f in files
        )
        messagebox.showinfo("完了", msg)
        for f in files:
            os.startfile(f)
    except Exception as e:
        messagebox.showerror("エラー", str(e))
    finally:
        root.destroy()
