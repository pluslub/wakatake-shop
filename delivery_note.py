import csv
import os
import sys
import shutil
from datetime import datetime
from tkinter import filedialog, messagebox

import openpyxl

TEMPLATE_NAME = 'ダイキン用納品書（総まとめ）.xlsx'
DATA_START_ROW = 7
DATA_END_ROW = 167


def get_template_path() -> str:
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, 'テンプレート', TEMPLATE_NAME)


def read_delivery_csv(filepath: str) -> list:
    rows = []
    with open(filepath, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        # BOM がヘッダーに残る場合に除去
        if reader.fieldnames:
            reader.fieldnames = [k.lstrip('﻿') for k in reader.fieldnames]
        for row in reader:
            name = row.get('商品名', '').strip()
            unit_price = row.get('単価(税込)', '0').strip().replace(',', '')
            qty = row.get('合計数量', '0').strip().replace(',', '')
            if not name:
                continue
            try:
                rows.append({
                    'name': name,
                    'unit_price': int(float(unit_price)),
                    'qty': int(float(qty)),
                })
            except ValueError:
                continue
    return rows


def fill_template(template_path: str, data: list, output_path: str) -> None:
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 納品日を今日に更新
    ws['C2'] = f'納品日　{datetime.now().strftime("%Y/%m/%d")}'

    # データ行をクリア（B〜E列のみ）
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws.cell(row=r, column=2).value = None
        ws.cell(row=r, column=3).value = None
        ws.cell(row=r, column=4).value = None
        ws.cell(row=r, column=5).value = None

    # CSVデータを書き込む
    for i, row in enumerate(data):
        r = DATA_START_ROW + i
        if r > DATA_END_ROW:
            break
        ws.cell(row=r, column=2).value = row['name']
        ws.cell(row=r, column=3).value = row['unit_price']
        ws.cell(row=r, column=4).value = row['qty']
        ws.cell(row=r, column=5).value = f'=C{r}*D{r}'

    wb.save(output_path)


def run() -> None:
    template_path = get_template_path()
    if not os.path.exists(template_path):
        messagebox.showerror('エラー', f'テンプレートが見つかりません:\n{template_path}')
        return

    csv_path = filedialog.askopenfilename(
        title='納品書用CSVを選択してください',
        filetypes=[('CSVファイル', '*.csv'), ('すべてのファイル', '*.*')],
    )
    if not csv_path:
        return

    try:
        data = read_delivery_csv(csv_path)
    except Exception as e:
        messagebox.showerror('エラー', f'CSV読み込みエラー:\n{e}')
        return

    if not data:
        dept_cols = {"部署名", "氏名", "注文番号"}
        try:
            with open(csv_path, encoding='utf-8-sig', newline='') as _f:
                import csv as _csv
                _reader = _csv.DictReader(_f)
                _cols = set(_reader.fieldnames or [])
        except Exception:
            _cols = set()
        if dept_cols.issubset(_cols):
            messagebox.showerror('エラー', '部署別CSVが選択されています。\nWordPressの「納品書用に出力」でダウンロードしたCSVを選択してください。')
        else:
            messagebox.showerror('エラー', 'CSVにデータがありません')
        return

    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_dir = os.path.dirname(csv_path)
    output_path = os.path.join(base_dir, f'納品書_{now}.xlsx')

    try:
        fill_template(template_path, data, output_path)
    except Exception as e:
        messagebox.showerror('エラー', f'Excel出力エラー:\n{e}')
        return

    messagebox.showinfo('完了', f'出力完了:\n{output_path}')
    os.startfile(output_path)


if __name__ == '__main__':
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    run()
    root.destroy()
