"""
令和8年夏季ダイキン注文書.xlsx の商品リスト部分（7〜220行目付近）を
今年の商品マスターCSV(20260630summer.csv)の131商品・7カテゴリーに
作り直すワンショットスクリプト。

- 1〜6行目（タイトル・部署名・氏名ヘッダー）はそのまま維持
- 7行目以降を全消去し、カテゴリーごとに1ブロックとして再構築
- 数量欄(D〜BF列、個人別)は空欄のまま（数量データは別途入力）
- 小計・個人別合計・総合計の数式も新しい行位置に合わせて再生成
"""
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

SRC_PATH = r'C:\Users\Workstation2402\Downloads\令和8年夏季ダイキン注文書.xlsx'
OUT_PATH = r'C:\Users\Workstation2402\Downloads\令和8年夏季ダイキン注文書_更新.xlsx'
MASTER_CSV = r'C:\Users\Workstation2402\Desktop\Documents\wakatake-shop\商品データ\20260630summer.csv'

CATEGORY_ORDER = [
    '若竹会',
    '味と暮らしの特選街',
    'カレーフェスタ',
    'お家で手軽に本格派カレー',
    'まぜごはん・玄米ごはん',
    'みそ汁・フリーズドライ',
    '健康茶',
]

D_COL = column_index_from_string('D')   # 4
BF_COL = column_index_from_string('BF')  # 58
BG_COL = column_index_from_string('BG')  # 59
BH_COL = column_index_from_string('BH')  # 60


def load_products():
    by_cat = {c: [] for c in CATEGORY_ORDER}
    with open(MASTER_CSV, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('名前', '').strip()
            cat = row.get('カテゴリー', '').strip()
            price = row.get('標準価格', '').strip()
            if cat in by_cat and name and price:
                by_cat[cat].append((name, int(float(price))))
    return by_cat


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=False)
    ws = wb['入力用']

    # 個人名リスト(D6:BF6)を先に控えておく(ブロックごとに再掲載するため)
    person_names = {}
    for c in range(D_COL, BF_COL + 1):
        person_names[c] = ws.cell(row=6, column=c).value

    # delete_rows()はセル結合の定義を正しく調整しないため、
    # 7行目以降にかかる結合セルを先に明示的に解除してから行を削除する
    stale_merges = [str(mc) for mc in ws.merged_cells.ranges if mc.min_row >= 7]
    for ref in stale_merges:
        ws.unmerge_cells(ref)

    # 7行目以降を全消去
    if ws.max_row >= 7:
        ws.delete_rows(7, ws.max_row - 6)

    by_cat = load_products()

    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    r = 7
    block_subtotal_rows = []

    for cat in CATEGORY_ORDER:
        items = by_cat[cat]
        if not items:
            continue

        # カテゴリー見出し行
        ws.cell(row=r, column=1, value=f'『　{cat}　』').font = header_font
        r += 1

        # サブヘッダー行(商品名/価格/お名前)
        ws.cell(row=r, column=2, value='商　品　名').font = header_font
        ws.cell(row=r, column=3, value='価格').font = header_font
        name_header_cell = ws.cell(row=r, column=D_COL, value='お　　名　　前')
        name_header_cell.font = header_font
        name_header_cell.alignment = center
        ws.merge_cells(start_row=r, start_column=D_COL, end_row=r, end_column=BF_COL)
        r += 1

        # 個人名の再掲載行
        for c in range(D_COL, BF_COL + 1):
            cell = ws.cell(row=r, column=c, value=person_names[c])
            cell.alignment = center
        r += 1

        block_start = r
        for i, (name, price) in enumerate(items, start=1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=price)
            ws.cell(row=r, column=BG_COL,
                    value=f'=SUM({get_column_letter(D_COL)}{r}:{get_column_letter(BF_COL)}{r})')
            ws.cell(row=r, column=BH_COL,
                    value=f'=SUMPRODUCT($C{r},{get_column_letter(BG_COL)}{r})')
            r += 1
        block_end = r - 1

        # 小計行
        subtotal_row = r
        ws.cell(row=subtotal_row, column=1, value='小　　　　　計').font = header_font
        for c in range(D_COL, BF_COL + 1):
            col_letter = get_column_letter(c)
            ws.cell(row=subtotal_row, column=c,
                    value=f'=SUMPRODUCT($C${block_start}:$C${block_end},'
                          f'{col_letter}{block_start}:{col_letter}{block_end})')
        ws.cell(row=subtotal_row, column=BG_COL,
                value=f'=SUM(BG{block_start}:BG{block_end})')
        ws.cell(row=subtotal_row, column=BH_COL,
                value=f'=SUM(BH{block_start}:BH{block_end})')
        block_subtotal_rows.append(subtotal_row)
        r = subtotal_row + 2  # 空行を1行挟んで次のブロックへ

    # 個人別合計
    r += 0
    name_row = r
    amount_row = r + 1
    ws.cell(row=name_row, column=1, value='個人別合計').font = header_font
    ws.cell(row=name_row, column=3, value='お名前').font = header_font
    ws.cell(row=amount_row, column=3, value='金額').font = header_font
    for c in range(D_COL, BF_COL + 1):
        col_letter = get_column_letter(c)
        ws.cell(row=name_row, column=c, value=f'={col_letter}6')
        sum_expr = '+'.join(f'{col_letter}{sr}' for sr in block_subtotal_rows)
        ws.cell(row=amount_row, column=c, value=f'={sum_expr}')

    # 総合計
    total_row = amount_row + 2
    ws.cell(row=total_row, column=1, value='総　合　計').font = header_font
    total_expr = '+'.join(f'BH{sr}' for sr in block_subtotal_rows)
    total_cell = ws.cell(row=total_row, column=D_COL, value=f'={total_expr}')
    total_cell.font = header_font

    wb.save(OUT_PATH)

    print(f'block_subtotal_rows = {block_subtotal_rows}')
    print(f'name_row={name_row} amount_row={amount_row} total_row={total_row}')
    print(f'saved: {OUT_PATH}')


if __name__ == '__main__':
    main()
