import csv
import os
from collections import OrderedDict
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


COLS = 5
COL_WIDTH_A = 16.88
COL_WIDTH_OTHER = 16.88
ROW_HEIGHT = 71.25
FONT_NAME = "游ゴシック"
FONT_SIZE = 11


def format_amount(total: float) -> str:
    n = int(round(total))
    formatted = f"{n:,}"
    table = str.maketrans("0123456789,", "０１２３４５６７８９，")
    return "￥" + formatted.translate(table)


def read_csv(filepath: str) -> OrderedDict:
    dept_data: OrderedDict = OrderedDict()
    with open(filepath, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [k.lstrip('﻿').strip() for k in reader.fieldnames]
        detected_cols = list(reader.fieldnames or [])
        for row in reader:
            dept = row.get("部署名", "").strip()
            name = row.get("氏名", "").strip()
            amount_str = row.get("合計金額(税込)", "0").strip().replace(",", "")
            if not dept:
                continue
            try:
                amount = float(amount_str)
            except ValueError:
                amount = 0.0
            if dept not in dept_data:
                dept_data[dept] = {"names": [], "total": 0.0}
            dept_data[dept]["total"] += amount
            if name and name not in dept_data[dept]["names"]:
                dept_data[dept]["names"].append(name)
    if not dept_data:
        delivery_cols = {"商品名", "単価(税込)", "合計数量"}
        if delivery_cols.issubset(set(detected_cols)):
            raise ValueError("納品書用CSVが選択されています。\nWordPressの「部署別に出力」でダウンロードしたCSVを選択してください。")
        raise ValueError(f"「部署名」列にデータが見つかりません\n検出された列: {detected_cols}")
    return dept_data


def make_thin_border() -> Border:
    thin = Side(border_style="thin", color="000000")
    return Border(top=thin, bottom=thin, left=thin, right=thin)


def apply_style(cell, border: Border, font_size: int = FONT_SIZE) -> None:
    cell.font = Font(name=FONT_NAME, size=font_size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def fill_sheet(ws, dept_data: OrderedDict, include_amount: bool,
               col_width: float = COL_WIDTH_A, row_height: float = ROW_HEIGHT,
               font_size: int = FONT_SIZE) -> None:
    border = make_thin_border()
    col_letters = [chr(ord("A") + i) for i in range(COLS)]

    for letter in col_letters:
        ws.column_dimensions[letter].width = col_width

    row_idx = 1
    col_idx = 1
    for dept, data in dept_data.items():
        lines = dept.split(" ", 1) if " " in dept else [dept]
        if include_amount:
            lines.append(format_amount(data["total"]))
        text = "\n".join(lines)

        cell = ws.cell(row=row_idx, column=col_idx, value=text)
        apply_style(cell, border, font_size)

        col_idx += 1
        if col_idx > COLS:
            col_idx = 1
            row_idx += 1

    max_row = row_idx if col_idx > 1 else row_idx - 1
    for i in range(1, max_row + 1):
        ws.row_dimensions[i].height = row_height

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"


def create_excel(dept_data: OrderedDict, output_path: str) -> None:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "集金用封筒（各部署へ配布）"
    fill_sheet(ws1, dept_data, include_amount=True)

    ws2 = wb.create_sheet("仕分け時ダンボール貼り付け用")
    fill_sheet(ws2, dept_data, include_amount=False,
               col_width=24.25, row_height=120, font_size=16)

    ws3 = wb.create_sheet("組合事務所での集金用")
    fill_sheet(ws3, dept_data, include_amount=True)

    wb.save(output_path)


def run() -> None:
    csv_path = filedialog.askopenfilename(
        title="CSVファイルを選択してください",
        filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")],
    )
    if not csv_path:
        return

    try:
        dept_data = read_csv(csv_path)
    except Exception as e:
        messagebox.showerror("エラー", f"CSV読み込みエラー:\n{e}")
        return

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = os.path.dirname(csv_path)
    output_path = os.path.join(base_dir, f"部署金額まとめ表_{now}.xlsx")

    try:
        create_excel(dept_data, output_path)
    except Exception as e:
        messagebox.showerror("エラー", f"Excel出力エラー:\n{e}")
        return

    messagebox.showinfo("完了", f"出力完了:\n{output_path}")
    os.startfile(output_path)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    run()
    root.destroy()
