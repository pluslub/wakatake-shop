"""Generate Excel order forms from PDF content."""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Workstation2402\Downloads\ilovepdf_extracted-pages (1)"
FN = 'MS Gothic'

def T(): return Side(style='thin')

def A():
    t = T()
    return Border(left=t, right=t, top=t, bottom=t)

def sc(ws, row, col, v='', bold=False, sz=9, ha='left', va='center',
       wrap=False, bd=None, bg=None):
    c = ws.cell(row=row, column=col)
    c.value = v
    c.font = Font(name=FN, bold=bold, size=sz)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bd is not None: c.border = bd
    if bg: c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    return c

def mc(ws, r1, c1, r2, c2, v='', bold=False, sz=9, ha='center', va='center',
       wrap=False, bd=None, bg=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return sc(ws, r1, c1, v, bold, sz, ha, va, wrap, bd, bg)

def cw(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

# ============================================================
# Union Service form (共通ヘッダー構造)
# cols: 1=No(4.5), 2=品名(20), 3=コード(6), 4=注文数(6),
#       5=sep(1.5), 6=No(4.5), 7=品名(20), 8=コード(6), 9=注文数(6)
# ============================================================
def build_union_form(title, items, fax_no='052-824-4152'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '発注書'
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    cw(ws, {1:4.5, 2:20, 3:6, 4:6, 5:1.5, 6:4.5, 7:20, 8:6, 9:6})

    # --- Title ---
    rh(ws, 1, 40); rh(ws, 2, 10)
    mc(ws, 1, 1, 2, 4, title, bold=True, sz=11, ha='left', va='center', wrap=True)
    mc(ws, 1, 6, 2, 9,
       '㈱ユニオンサービス\n本社事務所　名古屋市南区芝町189番地\n〒457-0023　TEL(052)824-4151',
       sz=7, ha='right', va='center', wrap=True)

    # --- 注文日 ---
    r = 3; rh(ws, r, 16)
    mc(ws, r, 1, r, 9, '注文日　　　　月　　　　日　（　　　　　　　）',
       sz=9, ha='left', bd=A())

    # --- お得意先 ---
    r = 4; rh(ws, r, 26)
    mc(ws, r, 1, r, 1, 'お得意先\nコードNO.', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 2, r, 3, '', bd=A())
    sc(ws, r, 4, '団体名', sz=8, ha='center', bd=A())
    mc(ws, r, 5, r, 8, '', bd=A())
    sc(ws, r, 9, '　様', sz=9, ha='right', bd=A())

    # --- 住所 ---
    r = 5; rh(ws, r, 16)
    sc(ws, r, 1, 'ご住所', sz=8, ha='center', bd=A())
    mc(ws, r, 2, r, 9, '〒', sz=9, ha='left', bd=A())

    r = 6; rh(ws, r, 22)
    sc(ws, r, 1, '', bd=A())
    mc(ws, r, 2, r, 4, '都道\n府県', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 5, r, 6, '市郡\n区', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 7, r, 9, '', bd=A())

    # --- 電話/FAX ---
    r = 7; rh(ws, r, 16)
    sc(ws, r, 1, '電話番号', sz=8, ha='center', bd=A())
    mc(ws, r, 2, r, 4, '', bd=A())
    mc(ws, r, 5, r, 6, 'FAX番号', sz=8, ha='center', bd=A())
    mc(ws, r, 7, r, 9, '', bd=A())

    # --- 担当者 ---
    r = 8; rh(ws, r, 22)
    sc(ws, r, 1, 'ご担当者様\n氏名', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 2, r, 4, '', bd=A())
    mc(ws, r, 5, r, 6, 'ご担当者様\n連絡先', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 7, r, 9, '', bd=A())

    # --- Table header ---
    r = 9; rh(ws, r, 22)
    sc(ws, r, 1, 'カタログ\nNo.', sz=8, ha='center', va='center', wrap=True, bd=A())
    sc(ws, r, 2, '品　　名', sz=9, ha='center', bd=A())
    sc(ws, r, 3, '商品\nコード', sz=8, ha='center', va='center', wrap=True, bd=A())
    sc(ws, r, 4, '注文数', sz=8, ha='center', va='center', wrap=True, bd=A())
    sc(ws, r, 6, 'カタログ\nNo.', sz=8, ha='center', va='center', wrap=True, bd=A())
    sc(ws, r, 7, '品　　名', sz=9, ha='center', bd=A())
    sc(ws, r, 8, '商品\nコード', sz=8, ha='center', va='center', wrap=True, bd=A())
    sc(ws, r, 9, '注文数', sz=8, ha='center', va='center', wrap=True, bd=A())

    # --- Items ---
    left  = [x for x in items if x[0] <= (len(items)//2 + len(items)%2)]
    right = [x for x in items if x[0] >  (len(items)//2 + len(items)%2)]
    nrows = max(len(left), len(right))
    sr = 10

    for i in range(nrows):
        r = sr + i; rh(ws, r, 15)
        if i < len(left):
            no, name, code = left[i]
            sc(ws, r, 1, no,   sz=8, ha='center', bd=A())
            sc(ws, r, 2, name, sz=8, ha='left', va='center', wrap=True, bd=A())
            sc(ws, r, 3, code, sz=8, ha='center', bd=A())
            sc(ws, r, 4, '',   bd=A())
        if i < len(right):
            no, name, code = right[i]
            sc(ws, r, 6, no,   sz=8, ha='center', bd=A())
            sc(ws, r, 7, name, sz=8, ha='left', va='center', wrap=True, bd=A())
            sc(ws, r, 8, code, sz=8, ha='center', bd=A())
            sc(ws, r, 9, '',   bd=A())

    # --- 小計 ---
    r = sr + nrows; rh(ws, r, 16)
    mc(ws, r, 1, r, 3, '小計', sz=9, ha='right', bd=A())
    sc(ws, r, 4, '', bd=A())
    mc(ws, r, 6, r, 8, '小計', sz=9, ha='right', bd=A())
    sc(ws, r, 9, '', bd=A())

    # --- 合計点数 ---
    r += 1; rh(ws, r, 18)
    mc(ws, r, 1, r, 8, '合計点数', bold=True, sz=10, ha='right', bd=A(), bg='CCCCCC')
    sc(ws, r, 9, '', bold=True, bd=A(), bg='CCCCCC')

    # --- 備考 ---
    r += 1
    mc(ws, r, 1, r, 9,
       '※お届け先、到着日などで指定の有る場合は、備考の欄に明記して下さい。', sz=8, ha='left')
    r += 1; rh(ws, r, 14)
    sc(ws, r, 1, '備考', sz=8, ha='center', bd=A())
    mc(ws, r, 2, r+3, 9, '', bd=A())
    for i in range(1, 4):
        rh(ws, r+i, 14)
        sc(ws, r+i, 1, '', bd=A())
    r += 4

    # --- FAX footer ---
    rh(ws, r, 30)
    mc(ws, r, 1, r, 2, '商品注文専用\nFAX', sz=8, ha='center', wrap=True, bd=A())
    mc(ws, r, 3, r, 9,
       f'0 5 2 ー 8 2 4 ー {fax_no.split("-")[-1]}',
       bold=True, sz=18, ha='center', bd=A())

    return wb


# ============================================================
# データ: 味と暮らしの特選街
# ============================================================
ITEMS_AJITO = [
    (1,  "黒わらび餅", "389"),
    (2,  "ホシフルーツ フルーツパーラーの凍らせて食べるシャーベット 彩果しずく4袋", "336"),
    (3,  "長野県産シャインマスカットひとくちゼリー", "342"),
    (4,  "和歌山 紀州南高梅 冷し中華", "83"),
    (5,  "戸隠 生そば(地粉)3人前", "84"),
    (6,  "来運たまごボーロ(おみくじ付き)", "349"),
    (7,  "愛媛 夕陽プレッツェル", "363"),
    (8,  "スティックフィナンシェ(瀬戸内レモン)", "362"),
    (9,  "レモネード 瀬戸田産レモン使用", "538"),
    (10, "水ようかん さらさS", "397"),
    (11, "竹炭豆 直送便", "398"),
    (12, "喜多方生ラーメン4食冷やし中華", "27"),
    (13, "小豆島 手延べそうめん 1.5kg", "34"),
    (14, "もっこす亭冷製冷しとまとラーメン", "4"),
    (15, "田舎造り 信州そば", "1"),
    (16, "高山ラーメン 鮎だし醤油味 4人前", "85"),
    (17, "ご当地ラーメン紀行", "79"),
    (18, "かけるごちそうシリーズ(3種セット)", "258"),
    (19, "さんわの手羽燻 9本詰め合せ(醤油・味噌・黒胡椒)", "287"),
    (20, "明太いわし＆いわし甘露煮詰め合せ", "208"),
    (21, "北海道産 ほたてめし", "293"),
    (22, "焼きたらチーズ", "228"),
    (23, "焼いか＆炙りチーズ", "206"),
    (24, "国産 味付うずらたまご", "239"),
    (25, "かつお節屋の万能つゆ 金の輝き・黒の輝き 2本セット", "259"),
    (26, "SABA 鰹ふりだし", "288"),
    (27, "食用オリーブ油", "268"),
    (28, "三種のお米せんべい", "330"),
    (29, "めちゃうまラー油あられ", "337"),
    (30, "シーサーのくるくるしっぽ シークワーサー風味", "332"),
    (31, "そら豆アソート(個包装テトラ)", "335"),
    (32, "トン 素焼きナッツ 食塩無添加ミックスナッツ", "216"),
    (33, "さかなっつハイ！ 10g×30P", "340"),
    (34, "炙り焼きあじ", "218"),
    (35, "健康菓集 ナッツ＆フルーツ", "202"),
    (36, "紀州南部の南高梅(はちみつ)", "290"),
    (37, "ちょび梅(3袋セット)", "317"),
    (38, "徳用 沖縄の塩黒糖(ピロ)", "301"),
    (39, "ドリップバッグ スペシャルブレンド", "552"),
    (40, "国産黒豆麦茶(2袋セット)", "550"),
    (41, "北海道コーンスープ", "158"),
    (42, "毎日のおみそ汁30P", "154"),
    (43, "いつものおみそ汁5種セット", "171"),
    (44, "野菜とたまごのスープ8P", "197"),
]

# ============================================================
# データ: カレーフェスタ
# ============================================================
ITEMS_CURRY = [
    (1,  "金沢 チャンピオンカレー(甘口)", "401"),
    (2,  "近江牛 牛すじスパイシービーフカレー", "402"),
    (3,  "瀬戸内えびのバターチキンカレー", "403"),
    (4,  "金沢 チャンピオンカレー(中辛)", "404"),
    (5,  "岐阜 柳家 飛騨牛カレー", "405"),
    (6,  "富良野スープカレー チキンレッグ1本入", "406"),
    (7,  "富良野スープカレー ほたて2個入り", "407"),
    (8,  "富良野スープカレー 厚切り豚バラ肉入り", "408"),
    (9,  "富良野ブラックカレー[ビーフ]", "409"),
    (10, "富良野バターチキンカレー", "410"),
    (11, "北海道マスカルポーネチーズカレー", "411"),
    (12, "前沢牛カレー「極旨」", "412"),
    (13, "利久 大きな牛たんカレー", "413"),
    (14, "三陸帆立ビーフカレー", "414"),
    (15, "タナカのおいしい若狭牛カレー", "415"),
    (16, "あふひ 費の極み 飛騨牛カレー(2人前)", "416"),
    (17, "飛騨牛 牛すじカレー", "417"),
    (18, "松阪牛ビーフカレー", "418"),
    (19, "奈良三条 ステーキ屋さんのビーフカレー", "419"),
    (20, "大阪・難波 自由軒名物カレー", "420"),
    (21, "堂島カレー ビーフオリジナル", "421"),
    (22, "神戸牛ビーフカレー", "422"),
    (23, "はかた地どりカレー", "423"),
    (24, "博多明太子カレー", "424"),
    (25, "博多華味鳥 ささみ入りチキンカレー", "425"),
    (26, "佐賀牛カレー プレミアム", "426"),
    (27, "かごしま黒豚カレー", "427"),
    (28, "あいがけカレー こだわりスパイスカレー×本格欧風ビーフカレー", "428"),
    (29, "あいがけカレー 和風スパイス×ココナッツチキンカレー", "429"),
    (30, "九州あいがけカレー 福岡×佐賀", "430"),
    (31, "右ちゃんまいう～!! もーもーカレービーフカレー", "431"),
    (32, "桃太郎カレー(甘口)", "432"),
    (33, "横須賀海軍カレー本舗 よこすか海軍カレー", "433"),
    (34, "桃太郎カレー", "434"),
    (35, "よこすか海軍カレー", "435"),
    (36, "トプカインド風ポークカレー(大辛)", "436"),
    (37, "スパイスボックス チキンカレー", "437"),
    (38, "ぶはらビーフカレー", "438"),
    (39, "ゲイロード バターチキンカレー", "439"),
    (40, "SPIKY きざみ生姜のキーマカレー", "440"),
    (41, "SPIKY ほぐし肉のスパイシーマサラ", "441"),
    (42, "SPIKY 実山椒香るチキンカレー", "442"),
    (43, "SPIKY 海老とトマトのレッドチリカレー", "443"),
    (44, "国産牛肉の100時間かけたビーフカレー 辛口", "444"),
    (45, "小野貫裕の鳥肌の立つカレー キーマカレー", "445"),
]


# ============================================================
# 若竹会 夏季物品販売申込書
# cols: 1=No(3), 2=商品名(18), 3=価格(6), 4-9=名前x6(7each), 10=数量(6), 11=金額(7)
# ============================================================
def build_wakatake():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '申込書'
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    cw(ws, {1:3, 2:18, 3:6, 4:7, 5:7, 6:7, 7:7, 8:7, 9:7, 10:6, 11:7})

    # Title
    r = 1; rh(ws, r, 22)
    mc(ws, r, 1, r, 11, '2026　【社会福祉法人　若竹会】　夏季物品販売申込書',
       bold=True, sz=13, ha='center')

    # 会社・団体名 / 代表者名
    r = 2; rh(ws, r, 18)
    mc(ws, r, 1, r, 5, '会社・団体名', sz=9, ha='left', bd=A())
    mc(ws, r, 6, r, 11, '代表者名', sz=9, ha='left', bd=A())

    r = 3; rh(ws, r, 6)  # spacer

    # Column headers (2 rows: お名前 spans cols 4-9 vertically)
    r = 4; rh(ws, r, 16)
    mc(ws, r, 1,  r+1, 1,  '',     bd=A())
    mc(ws, r, 2,  r+1, 2,  '商品名', sz=9, ha='center', bd=A())
    mc(ws, r, 3,  r+1, 3,  '価格',  sz=9, ha='center', bd=A())
    mc(ws, r, 4,  r,   9,  'お名前', sz=9, ha='center', bd=A())
    mc(ws, r, 10, r+1, 10, '数量',  sz=9, ha='center', bd=A())
    mc(ws, r, 11, r+1, 11, '金額',  sz=9, ha='center', bd=A())
    r = 5; rh(ws, r, 16)
    for col in range(4, 10):
        sc(ws, r, col, '', bd=A())

    def item_row(ws, r, no, name, price):
        rh(ws, r, 16)
        sc(ws, r, 1, no, sz=8, ha='center', bd=A())
        sc(ws, r, 2, name, sz=8, ha='left', bd=A())
        sc(ws, r, 3, price, sz=8, ha='right', bd=A())
        for col in range(4, 12):
            sc(ws, r, col, '', bd=A())

    def section_header(ws, r, text):
        rh(ws, r, 16)
        mc(ws, r, 1, r, 11, text, sz=9, ha='center', bd=A(), bg='EEEEEE')

    # Section 1
    r = 6; section_header(ws, r, "ベーカリー＆カフェ　わかたけ　『 焼き菓子 』")
    for no, name, price in [
        (1, "いちごのほっぺ（5個）", 250),
        (2, "天使のほっぺ（5個）", 200),
        (3, "キャラメルラスク（90g）", 400),
        (4, "シュガーラスク（80g）", 400),
        (5, "ガーリックラスク（75g）", 400),
        (6, "お菓子詰め合わせセット", 1400),
    ]:
        r += 1; item_row(ws, r, no, name, price)

    # Section 2
    r += 1; section_header(ws, r, "若竹作業所　『 おからクッキー 』")
    for no, name, price in [
        (7, "おからクッキー（きなこ）", 300),
        (8, "おからクッキー（プレーン）", 300),
    ]:
        r += 1; item_row(ws, r, no, name, price)

    # Section 3
    r += 1; section_header(ws, r, "山寺作業所　『 木製クリップ・エコたわし 』")
    for no, name, price in [
        (10, "手作り木製クリップ　4個入り", 150),
        (11, "エコたわし　2個入り", 100),
        (12, "エコたわし　10個入り", 400),
    ]:
        r += 1; item_row(ws, r, no, name, price)

    # 小計
    r += 1; rh(ws, r, 16)
    mc(ws, r, 1, r, 3, '小計', sz=9, ha='center', bd=A())
    for col in range(4, 12):
        sc(ws, r, col, '', bd=A())

    # 合計
    r += 1; rh(ws, r, 22)
    mc(ws, r, 1, r, 11, '合　計', bold=True, sz=11, ha='left', bd=A())

    # Footer
    r += 1; rh(ws, r, 14)
    mc(ws, r, 1, r, 6, '社会福祉法人　若竹会', sz=9, ha='left')
    mc(ws, r, 7, r, 11, 'TEL　077-569-5697', sz=9, ha='right')
    r += 1; rh(ws, r, 14)
    mc(ws, r, 7, r, 11, 'FAX　077-569-5518', sz=9, ha='right')

    return wb


# ============================================================
# F15 発注書
# cols: 1=商品名(20), 2=卸価格(9), 3=コード(6), 4=数量(6),
#       5=sep(1.5), 6=商品名(20), 7=卸価格(9), 8=コード(6), 9=数量(6)
# ============================================================
def build_f15():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '発注書'
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    cw(ws, {1:20, 2:9, 3:6.5, 4:6, 5:1.5, 6:20, 7:9, 8:6.5, 9:6})

    # Title
    r = 1; rh(ws, r, 30)
    mc(ws, r, 1, r, 9, '発注書　　FAX　052-821-8331', bold=True, sz=16, ha='left')

    # 注文日
    r = 2; rh(ws, r, 16)
    mc(ws, r, 5, r, 9, '注文日　　月　　日　（　　　）', sz=9, ha='right')

    # コードNO.
    r = 3; rh(ws, r, 16)
    sc(ws, r, 1, 'コードNO.', sz=8, ha='left', bd=A())
    mc(ws, r, 2, r, 4, '', bd=A())
    mc(ws, r, 5, r, 9,
       '※「注文日」「コードNO.」「ご担当名」のご記入をお願いします。', sz=8, ha='left')

    # 作業所名
    r = 4; rh(ws, r, 16)
    sc(ws, r, 1, '作業所名', sz=8, ha='left', bd=A())
    mc(ws, r, 2, r, 9, '', bd=A())

    # ご住所
    r = 5; rh(ws, r, 16)
    sc(ws, r, 1, 'ご住所　〒', sz=8, ha='left', bd=A())
    mc(ws, r, 2, r, 9, '', bd=A())

    r = 6; rh(ws, r, 22)
    mc(ws, r, 1, r, 4, '　　都道\n　　府県', sz=8, ha='left', wrap=True, bd=A())
    mc(ws, r, 5, r, 9, '', bd=A())

    r = 7; rh(ws, r, 6)  # spacer

    # TEL/FAX/担当
    r = 8; rh(ws, r, 16)
    sc(ws, r, 1, 'TEL', sz=8, ha='left', bd=A())
    sc(ws, r, 2, '', bd=A())
    sc(ws, r, 3, 'FAX', sz=8, ha='left', bd=A())
    sc(ws, r, 4, '', bd=A())
    sc(ws, r, 5, 'ご担当者', sz=8, ha='left', bd=A())
    sc(ws, r, 6, '', bd=A())
    mc(ws, r, 7, r, 7, 'ご担当者\n連絡先', sz=8, ha='left', wrap=True, bd=A())
    mc(ws, r, 8, r, 9, '', bd=A())

    # F15 section headers
    r = 9; rh(ws, r, 16)
    mc(ws, r, 1, r, 4, 'F15', bold=True, sz=11, ha='center', bd=A(), bg='DDDDDD')
    mc(ws, r, 6, r, 9, 'F15', bold=True, sz=11, ha='center', bd=A(), bg='DDDDDD')

    # Column headers
    r = 10; rh(ws, r, 22)
    sc(ws, r, 1, '商　品　名', sz=9, ha='center', bd=A())
    sc(ws, r, 2, '卸価格\n(税抜)', sz=8, ha='center', wrap=True, bd=A())
    sc(ws, r, 3, '商品\nコード', sz=8, ha='center', wrap=True, bd=A())
    sc(ws, r, 4, '数量', sz=8, ha='center', bd=A())
    sc(ws, r, 6, '商　品　名', sz=9, ha='center', bd=A())
    sc(ws, r, 7, '卸価格\n(税抜)', sz=8, ha='center', wrap=True, bd=A())
    sc(ws, r, 8, '商品\nコード', sz=8, ha='center', wrap=True, bd=A())
    sc(ws, r, 9, '数量', sz=8, ha='center', bd=A())

    left_items = [
        ("ビーフカレーインド風", "1,148円", "103"),
        ("ビーフカレーデラックスシルバーラベル", "1,148円", "105"),
        ("ココナッツミルク香るグリーンカレー【チキン】", "1,148円", "113"),
        ("スパイスにこだわったブラックカレー【ポーク】", "1,148円", "114"),
        ("ビーフ＆ベジタブルハイブレンド", "1,148円", "110"),
        ("ハヤシビーフ", "1,148円", "106"),
        ("カレーうどんソース金鯱", "1,148円", "122"),
        ("ひとくちカレー", "786円", "111"),
        ("減塩 ビーフカレー", "1,700円", "107"),
        ("減塩 野菜カレー", "1,700円", "108"),
        ("まゆまゆ Hokkori 玄米ごはんガバオライス", "1,447円", "142"),
        ("まゆまゆ Hokkori 玄米ごはん とりごはん", "1,447円", "143"),
        ("まゆまゆ Hokkori 玄米ごはん 五目カレーごはん", "1,447円", "144"),
        ("まぜこみご飯の素 五目ひじき", "1,133円", "145"),
        ("まぜこみご飯の素 牛ごぼう", "1,133円", "148"),
        ("まぜこみご飯の素 かしわめし", "1,133円", "141"),
    ]
    right_items = [
        ("いつものおみそ汁 贅沢 焼なす", "1,188円", "570"),
        ("いつものおみそ汁 贅沢 なめこ", "1,188円", "571"),
        ("いつものおみそ汁 贅沢 炒め野菜", "1,188円", "572"),
        ("いつものおみそ汁 贅沢 とうふ", "1,188円", "573"),
        ("いつものおみそ汁 贅沢 豚汁", "1,188円", "574"),
        ("海鮮雑炊4種セット4食", "747円", "575"),
        ("にゅうめん4種セット4食", "810円", "578"),
        ("減脂プーアール茶", "983円", "525"),
        ("ノンカフェイン ルイボスティー", "1,101円", "508"),
        ("杜仲茶", "1,022円", "514"),
        ("甜茶", "1,179円", "517"),
        ("ヘルシーどくだみ茶", "1,022円", "519"),
        ("はと麦茶", "1,337円", "522"),
        ("ジャスミン茶", "1,337円", "533"),
        ("(お徳用)煎茶ティーバッグ", "1,495円", "551"),
    ]

    sr = 11
    nrows = max(len(left_items), len(right_items))
    for i in range(nrows):
        r = sr + i; rh(ws, r, 15)
        if i < len(left_items):
            n, p, c = left_items[i]
            sc(ws, r, 1, n, sz=8, ha='left', wrap=True, bd=A())
            sc(ws, r, 2, p, sz=8, ha='right', bd=A())
            sc(ws, r, 3, c, sz=8, ha='center', bd=A())
            sc(ws, r, 4, '', bd=A())
        if i < len(right_items):
            n, p, c = right_items[i]
            sc(ws, r, 6, n, sz=8, ha='left', wrap=True, bd=A())
            sc(ws, r, 7, p, sz=8, ha='right', bd=A())
            sc(ws, r, 8, c, sz=8, ha='center', bd=A())
            sc(ws, r, 9, '', bd=A())

    # 合計数量
    r = sr + nrows; rh(ws, r, 18)
    mc(ws, r, 1, r, 5, '', bd=A())
    mc(ws, r, 6, r, 8, '合計数量', sz=9, ha='right', bd=A(), bg='CCCCCC')
    sc(ws, r, 9, '', bd=A(), bg='CCCCCC')

    # Note
    r += 1; rh(ws, r, 22)
    mc(ws, r, 1, r, 9,
       '※商品の注文金額合計が卸価格で10,000円(税抜)未満の場合は送料600円(税抜)が別途必要になります。',
       sz=7, ha='left', wrap=True)

    # Footer
    r += 1; rh(ws, r, 20)
    mc(ws, r, 1, r, 3, '㈱ユニオンサービス\n↑ FAX:052-821-8331', sz=8, ha='left', wrap=True)
    mc(ws, r, 4, r, 9,
       '本社事務所【〒457-0023 名古屋市南区芝町189番地／TEL052-822-2261】',
       sz=7, ha='right', wrap=True)

    return wb


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    os.makedirs(OUT, exist_ok=True)

    wb1 = build_union_form('2026年夏季\n味と暮らしの特選街専用発注書', ITEMS_AJITO)
    p1 = os.path.join(OUT, '注文書・料金表-2.xlsx')
    wb1.save(p1); print(f"OK: {p1}")

    wb2 = build_wakatake()
    p2 = os.path.join(OUT, '注文書・料金表-3.xlsx')
    wb2.save(p2); print(f"OK: {p2}")

    wb3 = build_union_form('カレーフェスタ専用発注書', ITEMS_CURRY)
    p3 = os.path.join(OUT, '注文書・料金表-5.xlsx')
    wb3.save(p3); print(f"OK: {p3}")

    wb4 = build_f15()
    p4 = os.path.join(OUT, '注文書・料金表-6.xlsx')
    wb4.save(p4); print(f"OK: {p4}")

    print("完了")
