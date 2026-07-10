"""
令和8年夏季ダイキン注文書.xlsx に、部署別CSVの注文データから
部署ごとの購入数量を書き込むワンショットスクリプト。

- 型番(A列)はこのスクリプトでは使わず、商品名(B列)でマッチングする
  (部署別CSVの商品内容には型番が含まれていないため)
- 部署名(部署別CSVの「部署名」列)をテンプレートのD〜BH列ヘッダー
  (3行目の部署名 + 6行目の氏名/ライン名を結合したもの)と突き合わせる
- マッチしなかった部署名・商品名は上書きせず、レポートに出す
"""
import csv
import re
import unicodedata
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

TEMPLATE_PATH = r'C:\Users\Workstation2402\Downloads\令和8年夏季ダイキン注文書.xlsx'
OUT_PATH = r'C:\Users\Workstation2402\Downloads\令和8年夏季ダイキン注文書_数量入力済み.xlsx'
DEPT_CSV = r'C:\Users\Workstation2402\Downloads\部署別_DKP+寺尾+正人_他52件_20260710.csv'

ITEM_RE = re.compile(r'^(.+)\[(.*)\]\s+x(\d+)$')
LEGACY_RE = re.compile(r'^(.+)\s+x(\d+)$')

D_COL = column_index_from_string('D')


def normalize(s: str) -> str:
    return ' '.join(unicodedata.normalize('NFKC', str(s)).split())


def parse_items(item_str: str):
    result = []
    for part in item_str.split(' | '):
        part = part.strip()
        m = ITEM_RE.match(part)
        if m:
            result.append((m.group(1).strip(), int(m.group(3))))
            continue
        m = LEGACY_RE.match(part)
        if m:
            result.append((m.group(1).strip(), int(m.group(2))))
    return result


def main():
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb['入力用']

    # 列マップ: 正規化した「部署名 氏名/ライン名」 -> 列番号
    col_map = {}
    for c in range(D_COL, ws.max_column + 1):
        dept = ws.cell(row=3, column=c).value
        name = ws.cell(row=6, column=c).value
        if name is None:
            continue
        combined = normalize(f'{dept} {name}') if dept else normalize(name)
        col_map[combined] = c

    # 行マップ: 正規化した商品名 -> 行番号 (型番が入っている行のみ対象)
    row_map = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None or b is None:
            continue
        # 見出し行(文字列で『』を含む)は除外、型番行(数字 or 'w-'始まり)のみ対象
        if isinstance(a, str) and a.startswith('『'):
            continue
        row_map[normalize(b)] = r

    # 部署別CSVを読み込んで集計: {列番号: {行番号: 合計数量}}
    matrix = {}
    unmatched_depts = set()
    unmatched_products = set()

    with open(DEPT_CSV, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [k.lstrip('﻿').strip() for k in reader.fieldnames]
        for row in reader:
            dept = row.get('部署名', '').strip()
            item_str = row.get('商品内容', '').strip()
            if not dept or not item_str:
                continue
            dept_norm = normalize(dept)
            col = col_map.get(dept_norm)
            if col is None:
                unmatched_depts.add(dept)
                continue
            for name, qty in parse_items(item_str):
                name_norm = normalize(name)
                r = row_map.get(name_norm)
                if r is None:
                    unmatched_products.add(name)
                    continue
                matrix.setdefault(col, {})
                matrix[col][r] = matrix[col].get(r, 0) + qty

    # 書き込み
    filled_cells = 0
    for col, row_qty in matrix.items():
        for r, qty in row_qty.items():
            ws.cell(row=r, column=col, value=qty)
            filled_cells += 1

    wb.save(OUT_PATH)

    print(f'書き込みセル数: {filled_cells}')
    print(f'マッチしなかった部署名: {len(unmatched_depts)}件')
    for d in sorted(unmatched_depts):
        print(f'  - {d}')
    print(f'マッチしなかった商品名: {len(unmatched_products)}件')
    for p in sorted(unmatched_products):
        print(f'  - {p}')
    print(f'保存先: {OUT_PATH}')


if __name__ == '__main__':
    main()
