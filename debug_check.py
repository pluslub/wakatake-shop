"""
発注書デバッグ確認スクリプト
注文書用CSVと出力済みExcelを照合し、数値の一致・不一致を報告します。
"""
import csv
import os
import sys
import unicodedata
import re
from tkinter import filedialog, Tk

import openpyxl

TEMPLATE_CONFIGS = {
    '味と暮らし': {'pairs': [('B','C','D'), ('G','H','I')], 'rows': (10, 31)},
    'カレーフェスタ': {'pairs': [('B','C','D'), ('G','H','I')], 'rows': (10, 32)},
    'F15':        {'pairs': [('A','C','D'), ('F','H','I')], 'rows': (11, 26)},
    '若竹会':     {'pairs': [('A',None,'I')],               'rows': (7,  19)},
}


def normalize(s: str) -> str:
    return ' '.join(unicodedata.normalize('NFKC', s).split())


def strip_item_number(s: str) -> str:
    return re.sub(r'^\d+[　\s]+', '', s)


def read_csv(filepath):
    qty_by_code = {}
    qty_by_name = {}
    with open(filepath, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('商品名', '').strip()
            code = row.get('商品コード', '').strip()
            qty_str = row.get('合計数量', '0').strip().replace(',', '')
            if not name:
                continue
            try:
                qty = int(float(qty_str))
            except ValueError:
                qty = 0
            if code:
                qty_by_code[code] = qty_by_code.get(code, 0) + qty
            qty_by_name[normalize(name)] = qty_by_name.get(normalize(name), 0) + qty
    return qty_by_code, qty_by_name


def read_excel_quantities(xlsx_path):
    """Excelから (商品名, コード, 書き込まれた数量) の一覧を返す"""
    fname = os.path.basename(xlsx_path)
    config = next((v for k, v in TEMPLATE_CONFIGS.items() if k in fname), None)
    if not config:
        return None, []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    min_row, max_row = config['rows']
    for name_col, code_col, qty_col in config['pairs']:
        for r in range(min_row, max_row + 1):
            name_val = ws[f'{name_col}{r}'].value
            if name_val is None:
                continue
            code_val = ws[f'{code_col}{r}'].value if code_col else None
            qty_val  = ws[f'{qty_col}{r}'].value
            rows.append({
                'name':  str(name_val).strip(),
                'code':  str(code_val).strip() if code_val else '',
                'qty':   int(qty_val) if qty_val is not None else None,
            })
    return fname, rows


def match_qty(row, qty_by_code, qty_by_name):
    """CSVから期待数量を返す（コード優先→名前照合）"""
    if row['code']:
        v = qty_by_code.get(row['code'])
        if v is not None:
            return v
    tmpl = normalize(strip_item_number(row['name']))
    v = qty_by_name.get(tmpl)
    if v is not None:
        return v
    tmpl_ns = tmpl.replace(' ', '')
    for csv_name, val in qty_by_name.items():
        if csv_name.replace(' ', '') == tmpl_ns:
            return val
    for csv_name, val in qty_by_name.items():
        if csv_name.startswith(tmpl) and len(csv_name) > len(tmpl) and csv_name[len(tmpl)] == ' ':
            return val
    return None


def check(csv_path, xlsx_paths):
    qty_by_code, qty_by_name = read_csv(csv_path)

    print(f"\nCSV: {os.path.basename(csv_path)}")
    print(f"対象商品数 (コード): {len(qty_by_code)}件 / (名前): {len(qty_by_name)}件\n")

    total_ok = total_ng = total_skip = 0

    for xlsx_path in xlsx_paths:
        fname, rows = read_excel_quantities(xlsx_path)
        if fname is None:
            print(f"[スキップ] {os.path.basename(xlsx_path)} （設定なし）\n")
            continue

        print(f"{'='*60}")
        print(f"【{fname}】")
        ok = ng = skip = 0
        for row in rows:
            expected = match_qty(row, qty_by_code, qty_by_name)
            actual   = row['qty']

            if expected is None and actual is None:
                skip += 1
                continue
            elif expected is None and actual is not None:
                print(f"  [!] 未注文なのに数量あり  : {row['name']} -> Excel={actual}")
                ng += 1
            elif expected is not None and actual is None:
                print(f"  [NG] 未入力              : {row['name']} (期待={expected})")
                ng += 1
            elif expected != actual:
                print(f"  [NG] 数量不一致          : {row['name']} (期待={expected}, Excel={actual})")
                ng += 1
            else:
                print(f"  [OK] {row['name']} = {actual}")
                ok += 1

        print(f"  -> 一致:{ok}件 / 不一致:{ng}件 / 注文なし:{skip}件\n")
        total_ok += ok; total_ng += ng; total_skip += skip

    print(f"{'='*60}")
    print(f"合計  [OK]{total_ok}件  [NG]{total_ng}件  スキップ:{total_skip}件")
    if total_ng == 0:
        print("-> すべて正常です！")
    else:
        print("-> 上記の不一致を確認してください。")


def main():
    root = Tk(); root.withdraw()

    csv_path = filedialog.askopenfilename(
        title='注文書用CSVを選択',
        filetypes=[('CSV', '*.csv'), ('すべて', '*.*')],
    )
    if not csv_path:
        return

    xlsx_paths = filedialog.askopenfilenames(
        title='確認するExcelファイルを選択（複数可）',
        filetypes=[('Excel', '*.xlsx'), ('すべて', '*.*')],
        initialdir=os.path.dirname(csv_path),
    )
    if not xlsx_paths:
        return

    check(csv_path, list(xlsx_paths))
    input("\nEnterキーで終了...")


if __name__ == '__main__':
    main()
