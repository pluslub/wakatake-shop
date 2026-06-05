"""
order_form_summer.py
注文書・料金表.pdf (7ページ) → Excel テンプレート生成
"""
import os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

FONT = "游ゴシック"
GRAY = "C0C0C0"
DARK_GRAY = "808080"


def _thin():
    return Side(border_style="thin", color="000000")

def _all_thin():
    t = _thin()
    return Border(top=t, bottom=t, left=t, right=t)

def _s(ws, row, col, value="", *, bold=False, sz=9, ha="center", va="center",
       wrap=True, bdr=True, bg=None, fc="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=sz, bold=bold, color=fc)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr:
        c.border = _all_thin()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _m(ws, r1, c1, r2, c2, value="", *, bold=False, sz=9, ha="center",
       va="center", wrap=True, bdr=True, bg=None, fc="000000"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(name=FONT, size=sz, bold=bold, color=fc)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr:
        c.border = _all_thin()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _rh(ws, row, height):
    ws.row_dimensions[row].height = height


# ── データ ──────────────────────────────────────────────────────────────────

# Page 1: 2026年夏季 味と暮らしの特選街 卸価格表 (No., 商品名, 販売価格税込, 卸価格税抜)
PRICE_SUMMER = [
    (1,  "黒わらび餅",                                                               1080, 850),
    (2,  "ホシフルーツ フルーツパーラーの凍らせて食べるシャーベット 彩果しずく4袋",  1510, 1188),
    (3,  "長野県産シャインマスカット ひとくちゼリー",                                 1080, 850),
    (4,  "和歌山 紀州南高梅 冷し中華",                                                860,  676),
    (5,  "戸隠 生そば(地粉)3人前",                                                    1290, 1014),
    (6,  "来運たまごボーロ(おみくじ付き)",                                            600,  471),
    (7,  "愛媛 夕陽プレッツェル",                                                     1080, 850),
    (8,  "スティックフィナンシェ〈瀬戸内レモン〉",                                    1620, 1275),
    (9,  "レモネード 瀬戸田産レモン使用",                                              1130, 889),
    (10, "水ようかん さらさ S",                                                        1430, 1125),
    (11, "竹炭豆 直送便",                                                              1180, 928),
    (12, "喜多方生ラーメン4食 冷やし中華",                                            1330, 1046),
    (13, "小豆島 手延べそうめん 1.5kg",                                                2260, 1778),
    (14, "もっこす亭冷製 冷しとまとラーメン",                                         1200, 944),
    (15, "田舎造り 信州そば",                                                          1580, 1242),
    (16, "高山ラーメン 鮎だし醤油味 4人前",                                            1080, 850),
    (17, "ご当地ラーメン紀行",                                                         1350, 1062),
    (18, "かけるごちそうシリーズ(3種セット)",                                         1220, 959),
    (19, "さんわの手羽煮 9本詰め合せ(醤油・味噌・黒胡椒)",                            1770, 1392),
    (20, "明太いわし&いわし甘露煮詰め合せ",                                           1380, 1085),
    (21, "北海道産 ほたてめし",                                                        1150, 904),
    (22, "焼きたらチーズ",                                                             1200, 944),
    (23, "焼いか&炙りチーズ",                                                         760,  597),
    (24, "国産 味付うずらたまご",                                                      850,  668),
    (25, "かつお節屋の万能つゆ 金の輝き・黒の輝き2本セット",                          1600, 1258),
    (26, "SABA 鰯ふりだし",                                                            1730, 1360),
    (27, "食用オリーブ油",                                                             1300, 1022),
    (28, "三種のお米せんべい",                                                         1060, 833),
    (29, "めちゃうまラー油あられ",                                                     750,  589),
    (30, "シーサーのくるくるしっぽ シークワーサー風味",                                970,  763),
    (31, "そら豆アソート(個包装テトラ)",                                               1350, 1062),
    (32, "トン 素焼きナッツ 食塩無添加ミックスナッツ",                                 1620, 1275),
    (33, "さかなっつハイ！ 10g×30P",                                                  1950, 1534),
    (34, "炙り焼きあじ",                                                               1300, 1022),
    (35, "健康菓集 ナッツ&フルーツ",                                                   1200, 944),
    (36, "紀州南部の南高梅(はちみつ)",                                                 1470, 1156),
    (37, "ちょび梅(3袋セット)",                                                        1080, 850),
    (38, "徳用 沖縄の塩黒糖(ピロ)",                                                    1250, 983),
    (39, "ドリップバッグ スペシャルブレンド",                                          2000, 1573),
    (40, "国産黒豆麦茶(2袋セット)",                                                    1200, 944),
    (41, "北海道コーンスープ",                                                         1190, 935),
    (42, "毎日のおみそ汁30P",                                                          1400, 1101),
    (43, "いつものおみそ汁5種セット",                                                  1180, 928),
    (44, "野菜とたまごのスープ8P",                                                     920,  723),
]

# Page 2: 若竹会 販売申込書
WAKATAKE_SECTIONS = [
    {
        "title": "ベーカリー＆カフェ　わかたけ　『　焼き菓子　』",
        "products": [
            (1,  "いちごのほっぺ（5個）",      250),
            (2,  "天使のほっぺ（5個）",        200),
            (3,  "キャラメルラスク（90g）",    400),
            (4,  "シュガーラスク（80g）",      400),
            (5,  "ガーリックラスク（75g）",    400),
            (6,  "お菓子詰め合わせセット",     1400),
        ],
    },
    {
        "title": "若竹作業所　『　おからクッキー』",
        "products": [
            (7,  "おからクッキー（きなこ）",   300),
            (8,  "おからクッキー（プレーン）",  300),
        ],
    },
    {
        "title": "山寺作業所　『　木製クリップ・エコたわし　』",
        "products": [
            (10, "手作り木製クリップ　4個入り", 150),
            (11, "エコたわし　2個入り",          100),
            (12, "エコたわし　10個入り",         400),
        ],
    },
]

# Page 3: 2026年夏季 発注書 (No., 商品名, 商品コード) — 22+22=44品
ORDER_SUMMER = [
    (1,  "黒わらび餅",                                                               389),
    (2,  "ホシフルーツ フルーツパーラーの凍らせて食べるシャーベット 彩果しずく4袋",  336),
    (3,  "長野県産シャインマスカットひとくちゼリー",                                  342),
    (4,  "和歌山 紀州南高梅 冷し中華",                                                83),
    (5,  "戸隠 生そば(地粉)3人前",                                                    84),
    (6,  "来運たまごボーロ(おみくじ付き)",                                            349),
    (7,  "愛媛 夕陽プレッツェル",                                                     363),
    (8,  "スティックフィナンシェ(瀬戸内レモン)",                                      362),
    (9,  "レモネード 瀬戸田産レモン使用",                                              538),
    (10, "水ようかん さらさ S",                                                        397),
    (11, "竹炭豆 直送便",                                                              398),
    (12, "喜多方生ラーメン4食冷やし中華",                                              27),
    (13, "小豆島 手延べそうめん 1.5kg",                                                34),
    (14, "もっこす亭冷製冷しとまとラーメン",                                           4),
    (15, "田舎造り 信州そば",                                                          1),
    (16, "高山ラーメン 鮎だし醤油味 4人前",                                            85),
    (17, "ご当地ラーメン紀行",                                                         79),
    (18, "かけるごちそうシリーズ(3種セット)",                                         258),
    (19, "さんわの手羽煮 9本詰め合せ(醤油・味噌・黒胡椒)",                            287),
    (20, "明太いわし&いわし甘露煮詰め合せ",                                           208),
    (21, "北海道産 ほたてめし",                                                        293),
    (22, "焼きたらチーズ",                                                             228),
    (23, "焼いか&炙りチーズ",                                                         206),
    (24, "国産 味付うずらたまご",                                                      239),
    (25, "かつお節屋の万能つゆ 金の輝き・黒の輝き2本セット",                          259),
    (26, "SABA 鰯ふりだし",                                                            288),
    (27, "食用オリーブ油",                                                             268),
    (28, "三種のお米せんべい",                                                         330),
    (29, "めちゃうまラー油あられ",                                                     337),
    (30, "シーサーのくるくるしっぽ シークワーサー風味",                                332),
    (31, "そら豆アソート(個包装テトラ)",                                               335),
    (32, "トン 素焼きナッツ 食塩無添加ミックスナッツ",                                 216),
    (33, "さかなっつハイ！ 10g×30P",                                                  340),
    (34, "炙り焼きあじ",                                                               218),
    (35, "健康菓集 ナッツ&フルーツ",                                                   202),
    (36, "紀州南部の南高梅(はちみつ)",                                                 290),
    (37, "ちょび梅(3袋セット)",                                                        317),
    (38, "徳用 沖縄の塩黒糖(ピロ)",                                                    301),
    (39, "ドリップバッグ スペシャルブレンド",                                          552),
    (40, "国産黒豆麦茶(2袋セット)",                                                    550),
    (41, "北海道コーンスープ",                                                         158),
    (42, "毎日のおみそ汁30P",                                                          154),
    (43, "いつものおみそ汁5種セット",                                                  171),
    (44, "野菜とたまごのスープ8P",                                                     197),
]

# Page 4: 2026年 カレーフェスタ 卸価格表 (No., 商品名, 販売価格税込, 卸価格税抜)
PRICE_CURRY = [
    (1,  "金沢 チャンピオンカレー(甘口)",                                  540,  401),
    (2,  "近江牛 牛すじスパイシービーフカレー",                            810,  597),
    (3,  "瀬戸内えびのバターチキンカレー",                                 540,  401),
    (4,  "金沢 チャンピオンカレー(中辛)",                                  540,  401),
    (5,  "岐阜 柳家 飛騨牛カレー",                                         885,  660),
    (6,  "富良野スープカレー チキンレッグ1本入",                           972,  723),
    (7,  "富良野スープカレー ほたて2個入り",                               864,  645),
    (8,  "富良野スープカレー 厚切り豚バラ肉入り",                          864,  645),
    (9,  "富良野ブラックカレー[ビーフ]",                                   529,  392),
    (10, "富良野バターチキンカレー",                                        561,  416),
    (11, "北海道マスカルポーネチーズカレー",                               777,  573),
    (12, "前沢牛カレー「極旨」",                                            864,  645),
    (13, "利久 大きな牛たんカレー",                                         1296, 1024),
    (14, "三陸帆立ビーフカレー",                                            756,  558),
    (15, "タナカのおいしい若狭牛カレー",                                    648,  479),
    (16, "あふひ 贅の極み 飛騨牛カレー(2人前)",                            1620, 1203),
    (17, "飛騨牛 牛すじカレー",                                             715,  527),
    (18, "松阪牛ビーフカレー",                                              648,  479),
    (19, "奈良三条ステーキ屋さんのビーフカレー",                           756,  558),
    (20, "大阪・難波 自由軒名物カレー",                                     564,  441),
    (21, "堂島カレー ビーフオリジナル",                                     1026, 763),
    (22, "神戸牛ビーフカレー",                                              648,  479),
    (23, "はかた地どりカレー",                                              518,  385),
    (24, "博多明太子カレー",                                                540,  424),
    (25, "博多華味鳥 ささみ入りチキンカレー",                               540,  401),
    (26, "佐賀牛カレー プレミアム",                                         1674, 1251),
    (27, "かごしま黒豚カレー",                                              518,  385),
    (28, "あいがけカレー こだわりスパイスカレー×本格欧風ビーフカレー",     669,  495),
    (29, "あいがけカレー 和風スパイス×ここナッツチキンカレー",             669,  495),
    (30, "九州あいがけカレー 福岡×佐賀",                                   702,  519),
    (31, "石ちゃんまいう〜!! もーもーカレービーフカレー",                  432,  322),
    (32, "桃太郎カレー(甘口)",                                              540,  401),
    (33, "横須賀海軍カレー本舗 よこすか海軍カレー",                         594,  440),
    (34, "桃太郎カレー",                                                    540,  401),
    (35, "よこすか海軍カレー",                                              523,  385),
    (36, "トプカインド風ポークカレー(大辛)",                                540,  401),
    (37, "スパイスボックス チキンカレー",                                   594,  440),
    (38, "ぶはらビーフカレー",                                              518,  385),
    (39, "ゲイロード バターチキンカレー",                                   518,  385),
    (40, "SPIKY きざみ生姜のキーマカレー",                                 421,  306),
    (41, "SPIKY ほぐし肉のスパイシーマサラ",                               421,  306),
    (42, "SPIKY 実山椒香るチキンカレー",                                    421,  306),
    (43, "SPIKY 海老とトマトのレッドチリカレー",                           421,  306),
    (44, "国産牛肉の100時間かけたビーフカレー辛口",                        486,  361),
    (45, "小野員裕の鳥肌の立つカレーキーマカレー",                          540,  401),
]

# Page 5: カレーフェスタ発注書 (No., 商品名, 商品コード) — 23+22=45品
ORDER_CURRY = [
    (1,  "金沢 チャンピオンカレー(甘口)",                                  401),
    (2,  "近江牛 牛すじスパイシービーフカレー",                            402),
    (3,  "瀬戸内えびのバターチキンカレー",                                 403),
    (4,  "金沢 チャンピオンカレー(中辛)",                                  404),
    (5,  "岐阜 柳家 飛騨牛カレー",                                         405),
    (6,  "富良野スープカレー チキンレッグ1本入",                           406),
    (7,  "富良野スープカレー ほたて2個入り",                               407),
    (8,  "富良野スープカレー 厚切り豚バラ肉入り",                          408),
    (9,  "富良野ブラックカレー[ビーフ]",                                   409),
    (10, "富良野バターチキンカレー",                                        410),
    (11, "北海道マスカルポーネチーズカレー",                               411),
    (12, "前沢牛カレー「極旨」",                                            412),
    (13, "利久 大きな牛たんカレー",                                         413),
    (14, "三陸帆立ビーフカレー",                                            414),
    (15, "タナカのおいしい若狭牛カレー",                                    415),
    (16, "あふひ 贅の極み 飛騨牛カレー(2人前)",                            416),
    (17, "飛騨牛 牛すじカレー",                                             417),
    (18, "松阪牛ビーフカレー",                                              418),
    (19, "奈良三条ステーキ屋さんのビーフカレー",                           419),
    (20, "大阪・難波 自由軒名物カレー",                                     420),
    (21, "堂島カレー ビーフオリジナル",                                     421),
    (22, "神戸牛ビーフカレー",                                              422),
    (23, "はかた地どりカレー",                                              423),
    (24, "博多明太子カレー",                                                424),
    (25, "博多華味鳥 ささみ入りチキンカレー",                               425),
    (26, "佐賀牛カレー プレミアム",                                         426),
    (27, "かごしま黒豚カレー",                                              427),
    (28, "あいがけカレー こだわりスパイスカレー×本格欧風ビーフカレー",     428),
    (29, "あいがけカレー 和風スパイス×ここナッツチキンカレー",             429),
    (30, "九州あいがけカレー 福岡×佐賀",                                   430),
    (31, "石ちゃんまいう〜!! もーもーカレービーフカレー",                  431),
    (32, "桃太郎カレー(甘口)",                                              432),
    (33, "横須賀海軍カレー本舗 よこすか海軍カレー",                         433),
    (34, "桃太郎カレー",                                                    434),
    (35, "よこすか海軍カレー",                                              435),
    (36, "トプカインド風ポークカレー(大辛)",                                436),
    (37, "スパイスボックス チキンカレー",                                   437),
    (38, "ぶはらビーフカレー",                                              438),
    (39, "ゲイロード バターチキンカレー",                                   439),
    (40, "SPIKY きざみ生姜のキーマカレー",                                 440),
    (41, "SPIKY ほぐし肉のスパイシーマサラ",                               441),
    (42, "SPIKY 実山椒香るチキンカレー",                                    442),
    (43, "SPIKY 海老とトマトのレッドチリカレー",                           443),
    (44, "国産牛肉の100時間かけたビーフカレー辛口",                        444),
    (45, "小野員裕の鳥肌の立つカレーキーマカレー",                          445),
]

# Page 6: F15 発注書（更新版）左列 (商品名, 卸価格税抜, 商品コード)
F15_NEW_LEFT = [
    ("ビーフカレーインド風",                              1148, 103),
    ("ビーフカレーデラックスシルバーラベル",              1148, 105),
    ("ココナッツミルク香るグリーンカレー【チキン】",      1148, 113),
    ("スパイスにこだわったブラックカレー【ポーク】",      1148, 114),
    ("ビーフ&ベジタブルハイブレンド",                    1148, 110),
    ("ハヤシビーフ",                                      1148, 106),
    ("カレーうどんソース金鯱",                            1148, 122),
    ("ひとくちカレー",                                    786,  111),
    ("減塩 ビーフカレー",                                 1700, 107),
    ("減塩 野菜カレー",                                   1700, 108),
    ("まゆまゆ Hokkori 玄米ごはんガバオライス",          1447, 142),
    ("まゆまゆ Hokkori 玄米ごはん とりごはん",           1447, 143),
    ("まゆまゆ Hokkori 玄米ごはん 五目カレーごはん",     1447, 144),
    ("まぜこみご飯の素 五目ひじき",                       1133, 145),
    ("まぜこみご飯の素 牛ごぼう",                         1133, 148),
    ("まぜこみご飯の素 かしわめし",                       1133, 141),
]

# Page 6: F15 発注書（更新版）右列
F15_NEW_RIGHT = [
    ("いつものおみそ汁 贅沢 焼なす",       1188, 570),
    ("いつものおみそ汁 贅沢 なめこ",       1188, 571),
    ("いつものおみそ汁 贅沢 炒め野菜",     1188, 572),
    ("いつものおみそ汁 贅沢 とうふ",       1188, 573),
    ("いつものおみそ汁 贅沢 豚汁",         1188, 574),
    ("海鮮雑炊4種セット4食",               747,  575),
    ("にゅうめん4種セット4食",             810,  578),
    ("減脂プーアール茶",                   983,  525),
    ("ノンカフェイン ルイボスティー",      1101, 508),
    ("杜仲茶",                             1022, 514),
    ("甜茶",                               1179, 517),
    ("ヘルシーどくだみ茶",                 1022, 519),
    ("はと麦茶",                           1337, 522),
    ("ジャスミン茶",                       1337, 533),
    ("(お徳用)煎茶ティーバッグ",           1495, 551),
]

# Page 7: F15 卸価格一覧表 (商品コード, 商品名, 販売価格税込, 卸価格税抜)
F15_PRICE_LIST = [
    (103, "ビーフカレーインド風",                                  1460, 1148),
    (105, "ビーフカレーデラックスシルバーラベル",                  1460, 1148),
    (113, "ココナッツミルク香るグリーンカレー【チキン】",           1460, 1148),
    (114, "スパイスにこだわったブラックカレー【ポーク】",           1460, 1148),
    (110, "ビーフ&ベジタブルハイブレンド",                         1460, 1148),
    (106, "ハヤシビーフ",                                           1460, 1148),
    (122, "カレーうどんソース金鯱",                                 1460, 1148),
    (111, "ひとくちカレー",                                         1000, 786),
    (107, "減塩 ビーフカレー",                                      2160, 1700),
    (108, "減塩 野菜カレー",                                        2160, 1700),
    (142, "まゆまゆ Hokkori 玄米ごはん　ガバオライス",             1840, 1447),
    (143, "まゆまゆ Hokkori 玄米ごはん　とりごはん",               1840, 1447),
    (144, "まゆまゆ Hokkori 玄米ごはん　五目カレーごはん",         1840, 1447),
    (145, "まぜこみご飯の素　五目ひじき",                           1440, 1133),
    (148, "まぜこみご飯の素　牛ごぼう",                             1440, 1133),
    (141, "まぜこみご飯の素　かしわめし",                           1440, 1133),
    (570, "いつものおみそ汁　贅沢　焼なす",                         1510, 1188),
    (571, "いつものおみそ汁　贅沢　なめこ",                         1510, 1188),
    (572, "いつものおみそ汁　贅沢　炒め野菜",                       1510, 1188),
    (573, "いつものおみそ汁　贅沢　とうふ",                         1510, 1188),
    (574, "いつものおみそ汁　贅沢　豚汁",                           1510, 1188),
    (575, "海鮮雑炊4種セット4食",                                   950,  747),
    (578, "にゅうめん4種セット4食",                                 1030, 810),
    (525, "減脂プーアール茶",                                        1250, 983),
    (508, "ノンカフェイン　ルイボスティー",                         1400, 1101),
    (514, "杜仲茶",                                                  1300, 1022),
    (517, "甜茶",                                                    1500, 1179),
    (519, "ヘルシーどくだみ茶",                                      1300, 1022),
    (522, "はと麦茶",                                                1700, 1337),
    (533, "ジャスミン茶",                                            1700, 1337),
    (551, "(お徳用) 煎茶ティーバッグ",                               1900, 1495),
]


# ── フォームビルダー ─────────────────────────────────────────────────────────

def _page_setup(ws, last_row, last_col="I"):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{last_col}{last_row}"


def create_price_list(ws, title, products, split, label="F15"):
    """2列の卸価格表（Pages 1・4）"""
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 28.0
    ws.column_dimensions["C"].width = 10.5
    ws.column_dimensions["D"].width = 10.5
    ws.column_dimensions["E"].width = 1.5
    ws.column_dimensions["F"].width = 5.0
    ws.column_dimensions["G"].width = 28.0
    ws.column_dimensions["H"].width = 10.5
    ws.column_dimensions["I"].width = 10.5

    _rh(ws, 1, 28)
    _m(ws, 1, 1, 1, 7, title, bold=True, sz=15, ha="center", va="center", bdr=False)
    _m(ws, 1, 8, 1, 9, label, bold=True, sz=11, ha="right", va="center", bdr=False)

    _rh(ws, 2, 22)
    for col, lbl in [(1, "No."), (2, "商　品　名"),
                     (3, "販売価格\n(税込)"), (4, "卸価格\n(税抜)")]:
        _s(ws, 2, col, lbl, bold=True, sz=8, bg=GRAY)
    _s(ws, 2, 5, "", bdr=False)
    for col, lbl in [(6, "No."), (7, "商　品　名"),
                     (8, "販売価格\n(税込)"), (9, "卸価格\n(税抜)")]:
        _s(ws, 2, col, lbl, bold=True, sz=8, bg=GRAY)

    left  = products[:split]
    right = products[split:]
    n     = max(len(left), len(right))

    for i in range(n):
        r = 3 + i
        _rh(ws, r, 28)
        if i < len(left):
            no, name, sell, ws_ = left[i]
            _s(ws, r, 1, no,              sz=9)
            _s(ws, r, 2, name,            sz=9, ha="left")
            _s(ws, r, 3, f"{sell:,}円",   sz=9)
            _s(ws, r, 4, f"{ws_:,}円",    sz=9)
        else:
            for c in range(1, 5):
                _s(ws, r, c, "")
        _s(ws, r, 5, "", bdr=False)
        if i < len(right):
            no, name, sell, ws_ = right[i]
            _s(ws, r, 6, no,              sz=9)
            _s(ws, r, 7, name,            sz=9, ha="left")
            _s(ws, r, 8, f"{sell:,}円",   sz=9)
            _s(ws, r, 9, f"{ws_:,}円",    sz=9)
        else:
            for c in range(6, 10):
                _s(ws, r, c, "")

    footer = 3 + n
    _rh(ws, footer, 20)
    _m(ws, footer, 1, footer, 9, "㈱ユニオンサービス", sz=10, ha="right", bdr=False)
    _page_setup(ws, footer)


def create_wakatake_form(ws):
    """Page 2: 若竹会 夏季物品販売申込書"""
    ws.title = "若竹会販売申込書"
    ws.column_dimensions["A"].width = 24.0
    ws.column_dimensions["B"].width = 8.0
    for letter in "CDEFGH":
        ws.column_dimensions[letter].width = 6.5
    ws.column_dimensions["I"].width = 7.0
    ws.column_dimensions["J"].width = 10.0
    NC = 10  # 総列数

    _rh(ws, 1, 26)
    _m(ws, 1, 1, 1, NC,
       "2026　【社会福祉法人　若竹会】　夏季物品販売申込書",
       bold=True, sz=13, ha="center", va="center", bdr=False)

    _rh(ws, 2, 22)
    _m(ws, 2, 1, 2, 5, "会社・団体名", sz=9, ha="left")
    _m(ws, 2, 6, 2, NC, "代表者名",    sz=9, ha="left")

    _rh(ws, 3, 8)
    _m(ws, 3, 1, 3, NC, "", bdr=False)

    # テーブルヘッダー（2行）
    _rh(ws, 4, 18)
    _rh(ws, 5, 16)
    _m(ws, 4, 1,  5, 1,  "商品名",  bold=True, sz=9, bg=GRAY)
    _m(ws, 4, 2,  5, 2,  "価格",    bold=True, sz=9, bg=GRAY)
    _m(ws, 4, 3,  4, 8,  "お名前",  bold=True, sz=9, bg=GRAY)
    for c in range(3, 9):
        _s(ws, 5, c, "", bg=GRAY)
    _m(ws, 4, 9,  5, 9,  "数量",    bold=True, sz=9, bg=GRAY)
    _m(ws, 4, 10, 5, 10, "金額",    bold=True, sz=9, bg=GRAY)

    row = 6
    for section in WAKATAKE_SECTIONS:
        _rh(ws, row, 22)
        _m(ws, row, 1, row, NC, section["title"],
           bold=True, sz=10, ha="center", bg=GRAY)
        row += 1
        for no, name, price in section["products"]:
            _rh(ws, row, 22)
            _s(ws, row, 1, f"{no}　{name}", sz=9, ha="left")
            _s(ws, row, 2, price, sz=9)
            for c in range(3, NC + 1):
                _s(ws, row, c, "")
            row += 1

    # 小計
    _rh(ws, row, 20)
    _m(ws, row, 1, row, 2, "小計", bold=True, sz=9, bg=GRAY)
    for c in range(3, NC + 1):
        _s(ws, row, c, "", bg=GRAY)
    row += 1

    _rh(ws, row, 8)
    _m(ws, row, 1, row, NC, "", bdr=False)
    row += 1

    # 合計
    _rh(ws, row, 26)
    _m(ws, row, 1, row, 2, "合　計", bold=True, sz=12)
    for c in range(3, NC + 1):
        _s(ws, row, c, "")
    row += 1

    _rh(ws, row, 18)
    _m(ws, row, 1, row, 5, "社会福祉法人　若竹会",                    sz=9, ha="left",  bdr=False)
    _m(ws, row, 6, row, NC, "TEL　077-569-5697　　FAX　077-569-5518", sz=9, ha="right", bdr=False)

    _page_setup(ws, row, "J")


def _header_standard(ws, line1, line2):
    """標準発注書ヘッダーを作成し、次の行番号を返す"""
    ws.column_dimensions["A"].width = 5.5
    ws.column_dimensions["B"].width = 24.0
    ws.column_dimensions["C"].width = 7.5
    ws.column_dimensions["D"].width = 7.0
    ws.column_dimensions["E"].width = 5.5
    ws.column_dimensions["F"].width = 24.0
    ws.column_dimensions["G"].width = 7.5
    ws.column_dimensions["H"].width = 7.0

    _rh(ws, 1, 24); _rh(ws, 2, 20)
    _m(ws, 1, 1, 2, 4, line1 + "\n" + line2,
       bold=True, sz=13, ha="left", va="center", bdr=False)
    _m(ws, 1, 5, 2, 8,
       "㈱ユニオンサービス\n本社事務所　名古屋市南区芝町189番地\n〒457-0023　TEL.(052)824-4151",
       sz=8, ha="right", va="center", bdr=False)

    _rh(ws, 3, 18)
    _m(ws, 3, 1, 3, 8, "注文日　　　月　　日　（　　　　　　　）", ha="left", va="center")

    _rh(ws, 4, 22)
    _m(ws, 4, 1, 4, 1, "お得意先\nコードNO.", sz=8)
    _m(ws, 4, 2, 4, 3, ""); _m(ws, 4, 4, 4, 4, "団体名", sz=8)
    _m(ws, 4, 5, 4, 7, ""); _s(ws, 4, 8, "様", sz=9)

    _rh(ws, 5, 18)
    _s(ws, 5, 1, "ご住所", sz=8); _s(ws, 5, 2, "〒", ha="left")
    _m(ws, 5, 3, 5, 4, ""); _m(ws, 5, 5, 5, 8, "")

    _rh(ws, 6, 18)
    _s(ws, 6, 1, ""); _m(ws, 6, 2, 6, 2, "都道\n府県", sz=7)
    _s(ws, 6, 3, ""); _m(ws, 6, 4, 6, 4, "市郡\n区", sz=7)
    _m(ws, 6, 5, 6, 8, "")

    _rh(ws, 7, 18)
    _s(ws, 7, 1, "電話番号", sz=8); _m(ws, 7, 2, 7, 3, "")
    _s(ws, 7, 4, ""); _s(ws, 7, 5, "FAX番号", sz=8)
    _m(ws, 7, 6, 7, 8, "")

    _rh(ws, 8, 18)
    _m(ws, 8, 1, 8, 1, "ご担当者様\n氏名",   sz=8)
    _m(ws, 8, 2, 8, 3, ""); _s(ws, 8, 4, "")
    _m(ws, 8, 5, 8, 5, "ご担当者様\n連絡先", sz=8)
    _m(ws, 8, 6, 8, 8, "")

    return 9


def _product_table(ws, start_row, products, split):
    """商品テーブル共通ロジック"""
    left  = products[:split]
    right = products[split:]
    n     = max(len(left), len(right))

    _rh(ws, start_row, 22)
    for col, lbl in [(1, "カタログ\nNo."), (2, "品　名"), (3, "商品\nコード"), (4, "注文数"),
                     (5, "カタログ\nNo."), (6, "品　名"), (7, "商品\nコード"), (8, "注文数")]:
        _s(ws, start_row, col, lbl, bold=True, sz=8, bg=GRAY)

    ds = start_row + 1
    for i in range(n):
        r = ds + i
        _rh(ws, r, 28)
        if i < len(left):
            no, name, code = left[i]
            _s(ws, r, 1, no, sz=9); _s(ws, r, 2, name, sz=9, ha="left")
            _s(ws, r, 3, code, sz=9); _s(ws, r, 4, "", sz=9)
        else:
            for c in range(1, 5): _s(ws, r, c, "")
        if i < len(right):
            no, name, code = right[i]
            _s(ws, r, 5, no, sz=9); _s(ws, r, 6, name, sz=9, ha="left")
            _s(ws, r, 7, code, sz=9); _s(ws, r, 8, "", sz=9)
        else:
            for c in range(5, 9): _s(ws, r, c, "")

    sub = ds + n
    _rh(ws, sub, 20)
    _m(ws, sub, 1, sub, 3, "小　計",   bold=True, sz=9, bg=GRAY)
    _s(ws, sub, 4, "", bg=GRAY)
    _m(ws, sub, 5, sub, 7, "合計点数", bold=True, sz=9, bg=DARK_GRAY, fc="FFFFFF")
    _s(ws, sub, 8, "", bg=DARK_GRAY)
    return sub + 1


def _footer_standard(ws, start_row, fax="052-824-4152"):
    r = start_row
    _rh(ws, r, 14)
    _m(ws, r, 1, r, 8,
       "※お届け先、到着日などで指定の有る場合は、備考の欄に明記して下さい。",
       sz=8, ha="left", bdr=False)
    r += 1
    _rh(ws, r, 16)
    _s(ws, r, 1, "備考", sz=8, ha="left")
    _m(ws, r, 2, r + 3, 8, "")
    for rr in range(r + 1, r + 4):
        _rh(ws, rr, 16); _s(ws, rr, 1, "")
    r += 4
    _rh(ws, r, 36)
    _m(ws, r, 1, r, 2, "商品注文専用\nFAX", sz=9, ha="left", va="center")
    _m(ws, r, 3, r, 8, fax, bold=True, sz=22, ha="center", va="center")
    _page_setup(ws, r)


def create_standard_order(ws, title1, title2, products, split, fax="052-824-4152"):
    """標準発注書（Pages 3・5）"""
    nr = _header_standard(ws, title1, title2)
    nr = _product_table(ws, nr, products, split)
    _footer_standard(ws, nr, fax)


def create_f15_new(ws):
    """Page 6: F15 発注書（更新版）"""
    ws.title = "F15発注書"
    ws.column_dimensions["A"].width = 24.0
    ws.column_dimensions["B"].width = 10.0
    ws.column_dimensions["C"].width = 7.0
    ws.column_dimensions["D"].width = 7.0
    ws.column_dimensions["E"].width = 24.0
    ws.column_dimensions["F"].width = 10.0
    ws.column_dimensions["G"].width = 7.0
    ws.column_dimensions["H"].width = 7.0

    _rh(ws, 1, 22); _rh(ws, 2, 18)
    _m(ws, 1, 1, 2, 4, "発注書", bold=True, sz=14, ha="left", va="center", bdr=False)
    _m(ws, 1, 5, 2, 8, "FAX　052－821－8331",
       bold=True, sz=16, ha="right", va="center", bdr=False)

    _rh(ws, 3, 18)
    _m(ws, 3, 5, 3, 8,
       "注文日　　　月　　日　（　　　　　）",
       ha="right", va="center", bdr=False)

    _rh(ws, 4, 18)
    _s(ws, 4, 1, "コードNO.", sz=8, ha="left")
    _m(ws, 4, 2, 4, 4, "")
    _m(ws, 4, 5, 4, 8,
       "※「注文日」「コードNO.」「ご担当名」のご記入をお願いします。",
       sz=8, ha="left", bdr=False)

    _rh(ws, 5, 18)
    _s(ws, 5, 1, "作業所名", sz=8, ha="left"); _m(ws, 5, 2, 5, 8, "")

    _rh(ws, 6, 18)
    _s(ws, 6, 1, "ご住所", sz=8, ha="left")
    _s(ws, 6, 2, "〒", ha="left"); _m(ws, 6, 3, 6, 8, "")

    _rh(ws, 7, 18)
    _s(ws, 7, 1, ""); _m(ws, 7, 2, 7, 2, "都道\n府県", sz=7)
    _m(ws, 7, 3, 7, 8, "")

    _rh(ws, 8, 18)
    _s(ws, 8, 1, "TEL", sz=9, ha="left"); _s(ws, 8, 2, "FAX", sz=9, ha="left")
    _m(ws, 8, 3, 8, 4, ""); _s(ws, 8, 5, "ご担当者", sz=8)
    _m(ws, 8, 6, 8, 6, "ご担当者\n連絡先", sz=8); _m(ws, 8, 7, 8, 8, "")

    # F15 区分見出し
    _rh(ws, 9, 20)
    _m(ws, 9, 1, 9, 4, "F15", bold=True, sz=12, bg=GRAY)
    _m(ws, 9, 5, 9, 8, "F15", bold=True, sz=12, bg=GRAY)

    # 商品ヘッダー
    _rh(ws, 10, 20)
    for col, lbl in [(1, "商　品　名"), (2, "卸価格\n(税抜)"),
                     (3, "商品\nコード"), (4, "数量"),
                     (5, "商　品　名"), (6, "卸価格\n(税抜)"),
                     (7, "商品\nコード"), (8, "数量")]:
        _s(ws, 10, col, lbl, bold=True, sz=8, bg=GRAY)

    n = max(len(F15_NEW_LEFT), len(F15_NEW_RIGHT))
    for i in range(n):
        r = 11 + i
        _rh(ws, r, 28)
        if i < len(F15_NEW_LEFT):
            name, price, code = F15_NEW_LEFT[i]
            _s(ws, r, 1, name,          sz=9, ha="left")
            _s(ws, r, 2, f"{price:,}円", sz=9)
            _s(ws, r, 3, code,          sz=9)
            _s(ws, r, 4, "",            sz=9)
        else:
            for c in range(1, 5): _s(ws, r, c, "")
        if i < len(F15_NEW_RIGHT):
            name, price, code = F15_NEW_RIGHT[i]
            _s(ws, r, 5, name,          sz=9, ha="left")
            _s(ws, r, 6, f"{price:,}円", sz=9)
            _s(ws, r, 7, code,          sz=9)
            _s(ws, r, 8, "",            sz=9)
        else:
            for c in range(5, 9): _s(ws, r, c, "")

    total_row = 11 + n
    _rh(ws, total_row, 20)
    _m(ws, total_row, 1, total_row, 7, "合計数量", bold=True, sz=9, bg=DARK_GRAY, fc="FFFFFF")
    _s(ws, total_row, 8, "", bg=DARK_GRAY)

    note_r = total_row + 1
    _rh(ws, note_r, 14)
    _m(ws, note_r, 1, note_r, 8,
       "※商品の注文金額合計が卸価格で10,000円(税抜)未満の場合は送料600円(税抜)が別途必要になります。",
       sz=8, ha="left", bdr=False)

    fax_r = note_r + 1
    _rh(ws, fax_r, 36)
    _m(ws, fax_r, 1, fax_r, 4,
       "㈱ユニオンサービス\nFAX:052-821-8331\n本社事務所〒457-0023名古屋市南区芝町189番地/TEL052-822-2261",
       sz=7, ha="left", va="center")
    _m(ws, fax_r, 5, fax_r, 8, "", bdr=False)

    _page_setup(ws, fax_r)


def create_f15_price_list(ws):
    """Page 7: F15 卸価格一覧表（単列）"""
    ws.title = "F15卸価格一覧表"
    ws.column_dimensions["A"].width = 9.0
    ws.column_dimensions["B"].width = 42.0
    ws.column_dimensions["C"].width = 13.0
    ws.column_dimensions["D"].width = 13.0

    _rh(ws, 1, 26)
    _m(ws, 1, 1, 1, 3, "卸価格一覧表", bold=True, sz=14, ha="left", va="center", bdr=False)
    _m(ws, 1, 4, 1, 4, "【F－15】",    bold=True, sz=12, ha="right", va="center", bdr=False)

    _rh(ws, 2, 22)
    _s(ws, 2, 1, "商品コード",       bold=True, sz=9, bg=GRAY)
    _s(ws, 2, 2, "商　品　名",        bold=True, sz=9, bg=GRAY)
    _m(ws, 2, 3, 2, 3, "販売価格\n(税込)", bold=True, sz=8, bg=GRAY)
    _m(ws, 2, 4, 2, 4, "卸価格\n(税抜)",   bold=True, sz=8, bg=GRAY)

    for i, (code, name, sell, ws_) in enumerate(F15_PRICE_LIST):
        r = 3 + i
        _rh(ws, r, 22)
        _s(ws, r, 1, code,          sz=9)
        _s(ws, r, 2, name,          sz=9, ha="left")
        _s(ws, r, 3, f"{sell:,}円", sz=9)
        _s(ws, r, 4, f"{ws_:,}円",  sz=9)

    last = 2 + len(F15_PRICE_LIST)
    _page_setup(ws, last, "D")


# ── エントリーポイント ────────────────────────────────────────────────────────

def get_template_dir():
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "テンプレート")


def run():
    tdir = get_template_dir()
    os.makedirs(tdir, exist_ok=True)
    outputs = []

    def save(filename, builder):
        wb = openpyxl.Workbook()
        builder(wb.active)
        path = os.path.join(tdir, filename)
        wb.save(path)
        outputs.append(path)

    # Page 1: 夏季 卸価格表
    def f1(ws):
        ws.title = "夏季卸価格表"
        create_price_list(ws, "2026年夏季　味と暮らしの特選街　卸価格表",
                          PRICE_SUMMER, split=22)
    save("卸価格表_夏季味と暮らし.xlsx", f1)

    # Page 2: 若竹会 販売申込書
    save("若竹会_夏季販売申込書.xlsx", create_wakatake_form)

    # Page 3: 夏季 発注書
    def f3(ws):
        ws.title = "夏季発注書"
        create_standard_order(ws,
            "2026年夏季",
            "味と暮らしの特選街専用発注書",
            ORDER_SUMMER, split=22)
    save("発注書_夏季味と暮らし.xlsx", f3)

    # Page 4: カレーフェスタ 卸価格表
    def f4(ws):
        ws.title = "カレーフェスタ卸価格表"
        create_price_list(ws, "2026年　カレーフェスタ企画　卸価格表",
                          PRICE_CURRY, split=23, label="(F15)")
    save("卸価格表_カレーフェスタ.xlsx", f4)

    # Page 5: カレーフェスタ 発注書
    def f5(ws):
        ws.title = "カレーフェスタ発注書"
        create_standard_order(ws,
            "カレーフェスタ専用発注書", "",
            ORDER_CURRY, split=23)
    save("発注書_カレーフェスタ.xlsx", f5)

    # Page 6: F15 発注書（更新版）
    save("発注書_F15夏季.xlsx", create_f15_new)

    # Page 7: F15 卸価格一覧表
    save("卸価格一覧表_F15.xlsx", create_f15_price_list)

    return outputs


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk(); root.withdraw()
    try:
        files = run()
        msg = "以下のファイルを生成しました:\n\n" + "\n".join(
            os.path.basename(f) for f in files)
        messagebox.showinfo("完了", msg)
        for f in files:
            os.startfile(f)
    except Exception as e:
        messagebox.showerror("エラー", str(e))
    finally:
        root.destroy()
