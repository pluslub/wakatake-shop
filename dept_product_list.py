import re
import csv
import os
import sys
from datetime import datetime
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CATEGORY_ORDER = [
    '若竹会',
    '味と暮らしの特選街',
    'カレーフェスタ',
    'お家で手軽に本格派カレー',
    'まぜごはん・玄米ごはん',
    'みそ汁・フリーズドライ',
    '健康茶',
]
UNCATEGORIZED = 'その他（未分類）'

ITEM_WITH_CATEGORY_RE = re.compile(r'^(.+)\[(.*)\]\s+x(\d+)$')
ITEM_LEGACY_RE = re.compile(r'^(.+)\s+x(\d+)$')


def load_category_map(master_csv_path: str) -> dict:
    """商品マスターCSV(名前・カテゴリー列)から {商品名: カテゴリー} を作成（旧形式CSV用フォールバック）"""
    name_to_category = {}
    with open(master_csv_path, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [k.lstrip('﻿').strip() for k in reader.fieldnames]
        for row in reader:
            name = row.get('名前', '').strip()
            category = row.get('カテゴリー', '').strip()
            if name and category:
                name_to_category[name] = category
    return name_to_category


def category_sort_key(category: str) -> tuple:
    if category in CATEGORY_ORDER:
        return (CATEGORY_ORDER.index(category), category)
    return (len(CATEGORY_ORDER), category)


def parse_items(item_str: str) -> list:
    """'商品名[カテゴリ] x数量 | ...' を [(商品名, カテゴリ, 数量), ...] に変換
    カテゴリ未埋め込みの旧形式（'商品名 x数量'）の場合はカテゴリ=Noneを返す"""
    result = []
    for part in item_str.split(' | '):
        part = part.strip()
        m = ITEM_WITH_CATEGORY_RE.match(part)
        if m:
            name = m.group(1).strip()
            category = m.group(2).strip() or None
            qty = int(m.group(3))
            result.append((name, category, qty))
            continue
        m = ITEM_LEGACY_RE.match(part)
        if m:
            result.append((m.group(1).strip(), None, int(m.group(2))))
    return result


def run() -> None:
    csv_path = filedialog.askopenfilename(
        title='部署別CSVを選択してください',
        filetypes=[('CSVファイル', '*.csv'), ('すべてのファイル', '*.*')],
    )
    if not csv_path:
        return

    dept_data = {}  # {部署名: {商品名: 数量}}
    name_to_category = {}  # CSVに埋め込まれたカテゴリーから構築

    try:
        with open(csv_path, encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            if reader.fieldnames:
                reader.fieldnames = [k.lstrip('﻿').strip() for k in reader.fieldnames]
            cols = set(reader.fieldnames or [])
            if '部署名' not in cols or '商品内容' not in cols:
                messagebox.showerror(
                    'エラー',
                    '部署別CSVを選択してください。\n'
                    'WordPressの「部署別CSVをダウンロード」で出力したファイルを選択してください。'
                )
                return
            for row in reader:
                dept = row.get('部署名', '').strip()
                item_str = row.get('商品内容', '').strip()
                if not dept or not item_str:
                    continue
                if dept not in dept_data:
                    dept_data[dept] = {}
                for name, category, qty in parse_items(item_str):
                    dept_data[dept][name] = dept_data[dept].get(name, 0) + qty
                    if category and name not in name_to_category:
                        name_to_category[name] = category
    except Exception as e:
        messagebox.showerror('エラー', f'CSV読み込みエラー:\n{e}')
        return

    if not dept_data:
        messagebox.showerror('エラー', 'CSVにデータがありません')
        return

    unmatched_names = set()
    use_categories = bool(name_to_category)

    # 旧形式CSV（カテゴリー未埋め込み）の場合は商品マスターCSVでの照合にフォールバック
    if not use_categories:
        master_csv_path = filedialog.askopenfilename(
            title='このCSVにはカテゴリー情報がありません。'
                  '商品マスターCSVを選択してください（キャンセルでカテゴリー分けなし）',
            filetypes=[('CSVファイル', '*.csv'), ('すべてのファイル', '*.*')],
        )
        if master_csv_path:
            try:
                name_to_category = load_category_map(master_csv_path)
                use_categories = bool(name_to_category)
            except Exception as e:
                messagebox.showerror('エラー', f'商品マスターCSV読み込みエラー:\n{e}')
                return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')
    category_font = Font(bold=True, color='1F2937')
    category_fill = PatternFill(fill_type='solid', fgColor='E5E7EB')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for dept in sorted(dept_data.keys()):
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', dept)[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells('A1:B1')
        ws['A1'].value = dept
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50

        ws['A2'].value = '商品名'
        ws['B2'].value = '数量'
        for cell in (ws['A2'], ws['B2']):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[2].height = 25

        grouped = {}  # {カテゴリー: [(商品名, 数量), ...]}
        for name, qty in dept_data[dept].items():
            category = name_to_category.get(name)
            if category is None:
                if use_categories:
                    unmatched_names.add(name)
                category = UNCATEGORIZED
            grouped.setdefault(category, []).append((name, qty))

        row = 3
        for category in sorted(grouped.keys(), key=category_sort_key):
            if use_categories:
                ws.merge_cells(f'A{row}:B{row}')
                cat_cell = ws.cell(row=row, column=1, value=category)
                cat_cell.font = category_font
                cat_cell.fill = category_fill
                cat_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[row].height = 22
                row += 1

            for name, qty in sorted(grouped[category]):
                c1 = ws.cell(row=row, column=1, value=name)
                c1.border = border
                c2 = ws.cell(row=row, column=2, value=qty)
                c2.alignment = Alignment(horizontal='center', vertical='center')
                c2.border = border
                ws.row_dimensions[row].height = 20
                row += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 8

    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(os.path.dirname(csv_path), f'部署別商品一覧_{now}.xlsx')

    try:
        wb.save(output_path)
    except Exception as e:
        messagebox.showerror('エラー', f'Excel出力エラー:\n{e}')
        return

    msg = f'出力完了:\n{output_path}'
    if unmatched_names:
        sample = '\n'.join(sorted(unmatched_names)[:10])
        more = f'\n...他{len(unmatched_names) - 10}件' if len(unmatched_names) > 10 else ''
        msg += (
            f'\n\n※カテゴリーが特定できず「{UNCATEGORIZED}」に分類された商品が'
            f'{len(unmatched_names)}件あります:\n{sample}{more}'
        )
    messagebox.showinfo('完了', msg)
    os.startfile(output_path)


if __name__ == '__main__':
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    run()
    root.destroy()
